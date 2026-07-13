

from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Request
from fastapi.responses import JSONResponse, Response
from sqlalchemy import text
import io
import heapq
import zipfile
from fastapi import FastAPI, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from starlette.concurrency import run_in_threadpool
import pandas as pd
import math
import re
import uuid
import hashlib
from collections import Counter
from pathlib import Path

from config import (
    CORS_ORIGINS,
    EVIDENCE_UPLOAD_DIR as EVIDENCE_UPLOAD_PATH,
    MAX_DATASET_UPLOAD_BYTES,
    MAX_EVIDENCE_UPLOAD_BYTES,
    MAX_EXCEL_UNCOMPRESSED_BYTES,
)
from database import engine, get_connection
from demo_seed import seed_demo_environment, seed_reference_catalog
from migrations import run_migrations
from routers.audit_log import router as audit_log_router
from routers.auth import router as auth_router
from routers.remediations import router as remediations_router
from security import bearer_token, user_from_token
from audit import write_audit

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=CORS_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

app.include_router(auth_router)
app.include_router(remediations_router)
app.include_router(audit_log_router)


PUBLIC_PATHS = {
    "/",
    "/health",
    "/api/auth/status",
    "/api/auth/bootstrap",
    "/api/auth/login",
    "/docs",
    "/openapi.json",
    "/redoc",
}

ALLOWED_EVIDENCE_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm", ".pdf", ".csv", ".txt"}
ALLOWED_DATASET_EXTENSIONS = {".xlsx", ".csv"}


def safe_client_filename(filename):
    original = Path(filename or "uploaded-file").name
    cleaned = re.sub(r"[^\w.\-\u0600-\u06ff ]+", "_", original, flags=re.UNICODE).strip(" ._")
    return cleaned[:180] or "uploaded-file"


def validate_upload_name(filename, allowed_extensions):
    original_name = safe_client_filename(filename)
    suffix = Path(original_name).suffix.lower()
    if suffix not in allowed_extensions:
        allowed = ", ".join(sorted(allowed_extensions))
        raise HTTPException(status_code=415, detail=f"Unsupported file type. Allowed: {allowed}")
    return original_name, suffix


async def read_upload_limited(file, allowed_extensions, max_upload_bytes=MAX_DATASET_UPLOAD_BYTES):
    original_name, suffix = validate_upload_name(file.filename, allowed_extensions)

    chunks = []
    total = 0
    while True:
        chunk = await file.read(1024 * 1024)
        if not chunk:
            break
        total += len(chunk)
        if total > max_upload_bytes:
            raise HTTPException(status_code=413, detail=f"File exceeds the {max_upload_bytes // (1024 * 1024)} MB limit")
        chunks.append(chunk)
    return original_name, suffix, b"".join(chunks)


def validate_excel_archive(file_path):
    try:
        with zipfile.ZipFile(file_path) as archive:
            entries = archive.infolist()
            uncompressed_size = sum(entry.file_size for entry in entries)
            compressed_size = sum(entry.compress_size for entry in entries)
    except zipfile.BadZipFile as exc:
        raise HTTPException(status_code=422, detail="The Excel file is not a valid XLSX archive") from exc

    compression_ratio = uncompressed_size / max(compressed_size, 1)
    if uncompressed_size > MAX_EXCEL_UNCOMPRESSED_BYTES or compression_ratio > 100:
        raise HTTPException(status_code=413, detail="The expanded Excel workbook exceeds the safe processing limit")


async def persist_evidence_upload(file):
    original_name, suffix = validate_upload_name(file.filename, ALLOWED_EVIDENCE_EXTENSIONS)
    stored_name = f"{uuid.uuid4().hex}{suffix}"
    destination = EVIDENCE_UPLOAD_PATH / stored_name
    EVIDENCE_UPLOAD_PATH.mkdir(parents=True, exist_ok=True)
    total = 0
    try:
        with destination.open("wb") as output:
            while True:
                chunk = await file.read(1024 * 1024)
                if not chunk:
                    break
                total += len(chunk)
                if total > MAX_EVIDENCE_UPLOAD_BYTES:
                    raise HTTPException(
                        status_code=413,
                        detail=f"File exceeds the {MAX_EVIDENCE_UPLOAD_BYTES // (1024 * 1024)} MB limit",
                    )
                output.write(chunk)
        if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
            validate_excel_archive(destination)
    except Exception:
        destination.unlink(missing_ok=True)
        raise
    return original_name, str(destination)


@app.middleware("http")
async def authentication_guard(request: Request, call_next):
    path = request.url.path
    if request.method == "OPTIONS" or path in PUBLIC_PATHS or path.startswith("/docs") or path.startswith("/redoc"):
        return await call_next(request)

    token = bearer_token(request)
    user = user_from_token(token)
    if not user:
        return JSONResponse(status_code=401, content={"detail": "Authentication required"})

    request.state.current_user = user
    if request.method in {"POST", "PUT", "PATCH", "DELETE"}:
        role = user["role"]
        reviewer_paths = ("/api/ndmo/answers", "/api/ndmo/ai-feedback", "/api/remediations")
        if role == "viewer" or (role == "reviewer" and not path.startswith(reviewer_paths)):
            return JSONResponse(status_code=403, content={"detail": "Insufficient permissions"})

    return await call_next(request)

@app.get("/")
def home():
    return {"message": "NDMO Platform Running"}


@app.get("/health")
def health():
    try:
        with engine.connect() as conn:
            conn.execute(text("SELECT 1"))
        return {
            "status": "healthy",
            "database": "connected",
            "analyzer": globals().get("AI_PROVIDER", "rule_based"),
        }
    except Exception as exc:
        raise HTTPException(status_code=503, detail="Database connection unavailable") from exc

@app.get("/questions")
def get_questions():

    with engine.connect() as conn:

        result = conn.execute(
            text("SELECT * FROM assessment_questions")
        )

        rows = result.fetchall()

        return [
            {
                "id": row[0],
                "question": row[1],
                "category": row[2],
                "weight": row[3]
            }
            for row in rows
        ]
        
        
@app.get("/score")
def calculate_score():

    with engine.connect() as conn:

        result = conn.execute(
            text("""
            SELECT
                q.category,
                COUNT(*) AS total_questions,
                SUM(CASE WHEN a.answer = TRUE THEN 1 ELSE 0 END) AS yes_answers
            FROM assessment_questions q
            JOIN assessment_answers a
                ON q.id = a.question_id
            GROUP BY q.category
            """)
        )

        rows = result.fetchall()

        categories = {}
        total_questions = 0
        total_yes = 0

        for row in rows:
            category = row.category
            total = row.total_questions
            yes = row.yes_answers

            score = (yes / total) * 100

            categories[category] = {
                "total_questions": total,
                "yes_answers": yes,
                "score": round(score, 2)
            }

            total_questions += total
            total_yes += yes

        ndi_score = (total_yes / total_questions) * 100

        return {
            "categories": categories,
            "overall": {
                "total_questions": total_questions,
                "yes_answers": total_yes,
                "ndi_score": round(ndi_score, 2)
            }
        }
        
@app.get("/recommendations")
def get_recommendations():

    with engine.connect() as conn:

        result = conn.execute(
            text("""
            SELECT
                q.category,
                q.question,
                a.answer
            FROM assessment_questions q
            JOIN assessment_answers a
                ON q.id = a.question_id
            WHERE a.answer = FALSE
            """)
        )

        rows = result.fetchall()

        recommendations = []

        for row in rows:
            if row.category == "Metadata":
                recommendations.append({
                    "category": row.category,
                    "issue": row.question,
                    "recommendation": "Create and maintain a Data Catalog that documents all data assets, owners, descriptions, and classifications."
                })

            elif row.category == "Sharing":
                recommendations.append({
                    "category": row.category,
                    "issue": row.question,
                    "recommendation": "Develop a Data Sharing Policy that defines approval process, sharing conditions, and responsible parties."
                })

            else:
                recommendations.append({
                    "category": row.category,
                    "issue": row.question,
                    "recommendation": "Review this requirement and prepare the needed evidence or policy."
                })

        return {
            "total_issues": len(recommendations),
            "recommendations": recommendations
        }
        
        
@app.get("/data-assets")
def get_data_assets():

    with engine.connect() as conn:

        result = conn.execute(
            text("""
            SELECT
                id,
                asset_name,
                description,
                owner_name,
                classification,
                records_count
            FROM data_assets
            ORDER BY id
            """)
        )

        rows = result.fetchall()

        return [
            {
                "id": row.id,
                "asset_name": row.asset_name,
                "description": row.description,
                "owner": row.owner_name,
                "classification": row.classification,
                "records_count": row.records_count
            }
            for row in rows
        ]
        
@app.get("/policies")
def get_policies():

    with engine.connect() as conn:

        result = conn.execute(
            text("""
            SELECT
                id,
                policy_name,
                policy_type,
                status,
                uploaded_at
            FROM policies
            ORDER BY id
            """)
        )

        rows = result.fetchall()

        return [
            {
                "id": row.id,
                "policy_name": row.policy_name,
                "policy_type": row.policy_type,
                "status": row.status,
                "uploaded_at": str(row.uploaded_at)
            }
            for row in rows
        ]
        
        
@app.get("/auto-score")
def auto_score():

    with engine.connect() as conn:

        # 1. Data Catalog exists?
        data_assets_count = conn.execute(
            text("SELECT COUNT(*) FROM data_assets")
        ).scalar()

        has_data_catalog = data_assets_count > 0

        # 2. All data assets have classification?
        missing_classification = conn.execute(
            text("""
            SELECT COUNT(*)
            FROM data_assets
            WHERE classification IS NULL OR classification = ''
            """)
        ).scalar()

        has_classification = missing_classification == 0 and data_assets_count > 0

        # 3. Data Governance Policy exists?
        governance_policy = conn.execute(
            text("""
            SELECT COUNT(*)
            FROM policies
            WHERE policy_type = 'Governance'
            AND status = 'Approved'
            """)
        ).scalar()

        has_governance_policy = governance_policy > 0

        # 4. Data Owner exists?
        assets_without_owner = conn.execute(
            text("""
            SELECT COUNT(*)
            FROM data_assets
            WHERE owner_name IS NULL OR owner_name = ''
            """)
        ).scalar()

        has_data_owner = assets_without_owner == 0 and data_assets_count > 0

        # 5. Data Quality Policy exists?
        quality_policy = conn.execute(
            text("""
            SELECT COUNT(*)
            FROM policies
            WHERE policy_type = 'Quality'
            AND status = 'Approved'
            """)
        ).scalar()

        has_quality_policy = quality_policy > 0

        # 6. Data Sharing Policy exists?
        sharing_policy = conn.execute(
            text("""
            SELECT COUNT(*)
            FROM policies
            WHERE policy_type = 'Sharing'
            AND status = 'Approved'
            """)
        ).scalar()

        has_sharing_policy = sharing_policy > 0

        checks = {
            "Governance": [
                has_governance_policy,
                has_data_owner
            ],
            "Metadata": [
                has_data_catalog
            ],
            "Classification": [
                has_classification
            ],
            "Quality": [
                has_quality_policy
            ],
            "Sharing": [
                has_sharing_policy
            ]
        }

        categories = {}
        total_checks = 0
        passed_checks = 0

        for category, values in checks.items():
            total = len(values)
            passed = sum(values)
            score = (passed / total) * 100

            categories[category] = {
                "total_checks": total,
                "passed_checks": passed,
                "score": round(score, 2)
            }

            total_checks += total
            passed_checks += passed

        ndi_score = (passed_checks / total_checks) * 100

        return {
            "categories": categories,
            "overall": {
                "total_checks": total_checks,
                "passed_checks": passed_checks,
                "ndi_score": round(ndi_score, 2)
            }
        }
        
        
@app.get("/auto-recommendations")
def auto_recommendations():
    with engine.connect() as conn:
        findings = conn.execute(text("""
            WITH latest_results AS (
                SELECT DISTINCT ON (question_id)
                    id, question_id, evidence_id, ai_answer, confidence, reason,
                    evidence_text, evidence_location, created_at
                FROM ndmo_ai_assessment_results
                ORDER BY question_id, created_at DESC, id DESC
            )
            SELECT
                r.id AS result_id,
                r.question_id,
                q.question_code,
                q.question_text_ar,
                q.question_text_en,
                r.evidence_id,
                e.file_name,
                e.evidence_type,
                r.ai_answer,
                r.confidence,
                r.reason,
                r.evidence_text,
                r.evidence_location,
                r.created_at,
                d.id AS domain_id,
                d.name_en AS domain_name_en,
                d.name_ar AS domain_name_ar,
                a.owner_name,
                a.due_date,
                a.status,
                a.notes,
                a.updated_at
            FROM latest_results r
            JOIN ndmo_questions q ON q.id = r.question_id
            JOIN ndmo_domains d ON d.id = q.domain_id
            JOIN ndmo_evidence e ON e.id = r.evidence_id
            LEFT JOIN remediation_actions a ON a.recommendation_key = 'ai-result-' || r.id::text
            WHERE LOWER(r.ai_answer) IN ('no', 'partial')
            ORDER BY
                CASE WHEN LOWER(r.ai_answer) = 'no' THEN 0 ELSE 1 END,
                r.confidence DESC,
                d.order_number,
                q.id
        """)).mappings().all()

    recommendations = []
    for row in findings:
        answer = str(row["ai_answer"] or "").lower()
        confidence = int(row["confidence"] or 0)
        priority = "Critical" if answer == "no" and confidence >= 70 else "High" if answer == "no" or confidence >= 70 else "Medium"
        recommendations.append({
            "id": f"ai-result-{row['result_id']}",
            "result_id": row["result_id"],
            "question_id": row["question_id"],
            "question_code": row["question_code"],
            "question_text_ar": row["question_text_ar"],
            "question_text_en": row["question_text_en"],
            "domain_id": row["domain_id"],
            "category": row["domain_name_en"],
            "category_ar": row["domain_name_ar"],
            "issue": f"Gap in requirement {row['question_code']}",
            "issue_ar": f"فجوة في المتطلب {row['question_code']}",
            "analysis": row["reason"] or "The uploaded evidence does not fully support this requirement.",
            "analysis_ar": row["reason"] or "الدليل المرفوع لا يدعم هذا المتطلب بشكل كامل.",
            "recommendation": "Review the cited evidence, correct the identified gap, and upload an updated approved file.",
            "recommendation_ar": "مراجعة الدليل المشار إليه ومعالجة الفجوة، ثم رفع ملف محدث ومعتمد.",
            "source": "Uploaded evidence assessment result",
            "evidence_id": row["evidence_id"],
            "evidence_file": row["file_name"],
            "evidence_type": row["evidence_type"],
            "evidence_location": row["evidence_location"],
            "evidence_text": row["evidence_text"],
            "ai_answer": answer,
            "confidence": confidence,
            "priority": priority,
            "owner_name": row["owner_name"],
            "due_date": row["due_date"],
            "status": row["status"] or "Open",
            "notes": row["notes"],
            "updated_at": row["updated_at"],
        })

    return {
        "total_issues": sum(1 for item in recommendations if item["status"] != "Resolved"),
        "critical_issues": sum(1 for item in recommendations if item["status"] != "Resolved" and item["priority"] == "Critical"),
        "recommendations": recommendations
    }
    
    
@app.post("/data-quality/upload")
async def upload_data_quality_file(
    file: UploadFile = File(...),
    asset_name: str = "Uploaded Dataset"
):

    contents = await file.read()
    file_name = file.filename.lower()

    try:
        if file_name.endswith(".csv"):
            df = pd.read_csv(io.BytesIO(contents))

        elif file_name.endswith(".xlsx"):
            df = pd.read_excel(
                io.BytesIO(contents),
                engine="openpyxl"
            )

        else:
            return {
                "error": "Unsupported file type. Please upload CSV or XLSX file only."
            }

    except Exception as e:
        return {
            "error": "File reading failed",
            "details": str(e)
        }

    total_rows = len(df)
    total_columns = len(df.columns)

    missing_values = int(df.isna().sum().sum())
    duplicate_rows = int(df.duplicated().sum())

    total_cells = total_rows * total_columns

    if total_cells > 0:
        completeness_score = (1 - (missing_values / total_cells)) * 100
    else:
        completeness_score = 0

    if total_rows > 0:
        uniqueness_score = (1 - (duplicate_rows / total_rows)) * 100
    else:
        uniqueness_score = 0

    quality_score = (completeness_score + uniqueness_score) / 2

    with engine.begin() as conn:
        conn.execute(
            text("""
            INSERT INTO data_quality_reports
            (
                asset_name,
                file_name,
                total_rows,
                total_columns,
                missing_values,
                duplicate_rows,
                completeness_score,
                uniqueness_score,
                quality_score
            )
            VALUES
            (
                :asset_name,
                :file_name,
                :total_rows,
                :total_columns,
                :missing_values,
                :duplicate_rows,
                :completeness_score,
                :uniqueness_score,
                :quality_score
            )
            """),
            {
                "asset_name": asset_name,
                "file_name": file.filename,
                "total_rows": total_rows,
                "total_columns": total_columns,
                "missing_values": missing_values,
                "duplicate_rows": duplicate_rows,
                "completeness_score": round(completeness_score, 2),
                "uniqueness_score": round(uniqueness_score, 2),
                "quality_score": round(quality_score, 2)
            }
        )

    return {
        "asset_name": asset_name,
        "file_name": file.filename,
        "total_rows": total_rows,
        "total_columns": total_columns,
        "missing_values": missing_values,
        "duplicate_rows": duplicate_rows,
        "completeness_score": round(completeness_score, 2),
        "uniqueness_score": round(uniqueness_score, 2),
        "quality_score": round(quality_score, 2)
    }
    
@app.get("/data-quality/reports")
def get_data_quality_reports():

    with engine.connect() as conn:

        result = conn.execute(
            text("""
            SELECT
                id,
                asset_name,
                file_name,
                total_rows,
                total_columns,
                missing_values,
                duplicate_rows,
                completeness_score,
                uniqueness_score,
                validity_score,
                quality_score,
                created_at
            FROM data_quality_reports
            ORDER BY id DESC
            """)
        )

        rows = result.fetchall()

        return [
            {
                "id": row.id,
                "asset_name": row.asset_name,
                "file_name": row.file_name,
                "total_rows": row.total_rows,
                "total_columns": row.total_columns,
                "missing_values": row.missing_values,
                "duplicate_rows": row.duplicate_rows,
                "completeness_score": float(row.completeness_score),
                "uniqueness_score": float(row.uniqueness_score),
                "validity_score": float(row.validity_score),
                "quality_score": float(row.quality_score),
                "created_at": str(row.created_at)
            }
            for row in rows
        ]
        
@app.get("/data-quality/recommendations")
def get_data_quality_recommendations():

    with engine.connect() as conn:

        result = conn.execute(
            text("""
            SELECT
                id,
                asset_name,
                file_name,
                total_rows,
                total_columns,
                missing_values,
                duplicate_rows,
                completeness_score,
                uniqueness_score,
                quality_score,
                created_at
            FROM data_quality_reports
            ORDER BY id DESC
            LIMIT 1
            """)
        )

        row = result.fetchone()

        if row is None:
            return {
                "message": "No data quality reports found.",
                "recommendations": []
            }

        recommendations = []

        if row.missing_values > 0:
            recommendations.append({
                "issue": "Missing values detected",
                "details": f"The dataset contains {row.missing_values} missing values.",
                "recommendation": "Review empty cells and complete missing required fields before using the dataset."
            })

        if row.duplicate_rows > 0:
            recommendations.append({
                "issue": "Duplicate rows detected",
                "details": f"The dataset contains {row.duplicate_rows} duplicate rows.",
                "recommendation": "Remove duplicate records to improve data uniqueness and reliability."
            })

        if float(row.completeness_score) < 80:
            recommendations.append({
                "issue": "Low completeness score",
                "details": f"Completeness score is {float(row.completeness_score)}%.",
                "recommendation": "Improve data collection process and make important fields mandatory."
            })

        if float(row.uniqueness_score) < 90:
            recommendations.append({
                "issue": "Low uniqueness score",
                "details": f"Uniqueness score is {float(row.uniqueness_score)}%.",
                "recommendation": "Apply duplicate detection rules and define unique identifiers for records."
            })

        if float(row.quality_score) >= 90:
            quality_level = "High"
        elif float(row.quality_score) >= 70:
            quality_level = "Medium"
        else:
            quality_level = "Low"

        return {
            "report_id": row.id,
            "asset_name": row.asset_name,
            "file_name": row.file_name,
            "quality_score": float(row.quality_score),
            "quality_level": quality_level,
            "total_recommendations": len(recommendations),
            "recommendations": recommendations
        }
        
        
latest_report = None


def get_quality_level(score):
    if score >= 90:
        return "High"
    elif score >= 75:
        return "Medium"
    elif score >= 60:
        return "Low"
    else:
        return "Very Low"


def build_recommendations(missing_values, duplicate_rows, completeness_score, uniqueness_score):
    recommendations = []

    if missing_values > 0:
        recommendations.append({
            "issue": "Missing values detected",
            "details": f"The dataset contains {missing_values} missing values.",
            "recommendation": "Review empty cells and complete missing required fields before using the dataset."
        })

    if duplicate_rows > 0:
        recommendations.append({
            "issue": "Duplicate rows detected",
            "details": f"The dataset contains {duplicate_rows} duplicate rows.",
            "recommendation": "Remove duplicate records to improve data uniqueness and reliability."
        })

    if completeness_score < 80:
        recommendations.append({
            "issue": "Low completeness score",
            "details": f"Completeness score is {round(completeness_score, 2)}%.",
            "recommendation": "Improve data collection process and make important fields mandatory."
        })

    if uniqueness_score < 90:
        recommendations.append({
            "issue": "Low uniqueness score",
            "details": f"Uniqueness score is {round(uniqueness_score, 2)}%.",
            "recommendation": "Apply duplicate detection rules and define unique identifiers for records."
        })

    if not recommendations:
        recommendations.append({
            "issue": "Good data quality",
            "details": "No major data quality issues were detected.",
            "recommendation": "Continue monitoring data quality regularly."
        })

    return recommendations


def parse_quality_rules(rules_json):
    try:
        raw = json.loads(rules_json or "{}")
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=422, detail="Quality rules must be valid JSON") from exc

    if not isinstance(raw, dict):
        raise HTTPException(status_code=422, detail="Quality rules must be a JSON object")
    return {
        "required_columns": [str(value).strip() for value in raw.get("required_columns", []) if str(value).strip()],
        "unique_columns": [str(value).strip() for value in raw.get("unique_columns", []) if str(value).strip()],
        "date_columns": [str(value).strip() for value in raw.get("date_columns", []) if str(value).strip()],
    }


def detect_pii_type(column_name, series):
    normalized_name = str(column_name).strip().lower().replace("_", " ")
    if any(term in normalized_name for term in ["email", "e mail", "البريد", "ايميل"]):
        return "email"
    if any(term in normalized_name for term in ["phone", "mobile", "telephone", "جوال", "هاتف"]):
        return "phone"
    if any(term in normalized_name for term in ["national id", "identity", "هوية", "اقامة", "إقامة"]):
        return "national_id"
    if any(term in normalized_name for term in ["full name", "customer name", "employee name", "الاسم", "اسم العميل"]):
        return "name"

    samples = series.dropna().astype(str).head(80)
    if not samples.empty and (samples.str.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$").mean() >= 0.8):
        return "email"
    return None


def profile_dataframe(df, rules):
    profiles = []
    issues = []
    issue_limit = 20_000
    required_columns = set(rules["required_columns"])
    unique_columns = set(rules["unique_columns"])
    date_columns = set(rules["date_columns"])

    for missing_column in sorted(required_columns - set(map(str, df.columns))):
        issues.append({"column_name": missing_column, "row_number": None, "issue_type": "missing_required_column", "issue_value": ""})

    for column in df.columns:
        column_name = str(column)
        series = df[column]
        total_values = len(series)
        missing_mask = series.isna()
        missing_values = int(missing_mask.sum())
        unique_values = int(series.nunique(dropna=True))
        duplicate_values = int(series.duplicated(keep=False).sum() - missing_values)
        pii_type = detect_pii_type(column_name, series)
        invalid_mask = pd.Series(False, index=series.index)

        if column_name in date_columns:
            parsed_dates = pd.to_datetime(series, errors="coerce")
            invalid_mask = series.notna() & parsed_dates.isna()
        elif pii_type == "email":
            invalid_mask = series.notna() & ~series.astype(str).str.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
        elif pii_type == "phone":
            digits = series.astype(str).str.replace(r"\D", "", regex=True)
            invalid_mask = series.notna() & ~digits.str.len().between(8, 15)

        valid_non_missing = max(total_values - missing_values, 0)
        invalid_values = int(invalid_mask.sum())
        validity_score = 100 if valid_non_missing == 0 else round((1 - invalid_values / valid_non_missing) * 100, 2)

        profiles.append({
            "column_name": column_name,
            "data_type": str(series.dtype),
            "total_values": total_values,
            "missing_values": missing_values,
            "unique_values": unique_values,
            "duplicate_values": max(duplicate_values, 0),
            "validity_score": validity_score,
            "pii_type": pii_type,
        })

        if len(issues) < issue_limit:
            for position, value in enumerate(missing_mask.tolist(), start=2):
                if value:
                    issues.append({"column_name": column_name, "row_number": position, "issue_type": "missing_value", "issue_value": ""})
                    if len(issues) >= issue_limit:
                        break

        if column_name in unique_columns and len(issues) < issue_limit:
            duplicate_mask = series.notna() & series.duplicated(keep=False)
            for position, (is_duplicate, value) in enumerate(zip(duplicate_mask.tolist(), series.tolist()), start=2):
                if is_duplicate:
                    issues.append({"column_name": column_name, "row_number": position, "issue_type": "duplicate_unique_value", "issue_value": str(value)[:500]})
                    if len(issues) >= issue_limit:
                        break

        if invalid_values and len(issues) < issue_limit:
            for position, (is_invalid, value) in enumerate(zip(invalid_mask.tolist(), series.tolist()), start=2):
                if is_invalid:
                    issues.append({"column_name": column_name, "row_number": position, "issue_type": "invalid_format", "issue_value": str(value)[:500]})
                    if len(issues) >= issue_limit:
                        break

    if len(issues) < issue_limit:
        duplicate_rows_mask = df.duplicated(keep=False)
        for position, is_duplicate in enumerate(duplicate_rows_mask.tolist(), start=2):
            if is_duplicate:
                issues.append({"column_name": None, "row_number": position, "issue_type": "duplicate_row", "issue_value": ""})
                if len(issues) >= issue_limit:
                    break

    validity_scores = [profile["validity_score"] for profile in profiles]
    if required_columns - set(map(str, df.columns)):
        validity_scores.extend([0] * len(required_columns - set(map(str, df.columns))))
    overall_validity = round(sum(validity_scores) / len(validity_scores), 2) if validity_scores else 100
    return profiles, issues, overall_validity


def load_quality_report(report_id=None):
    conn = get_connection()
    cur = conn.cursor()
    if report_id is None:
        cur.execute("SELECT * FROM data_quality_reports ORDER BY id DESC LIMIT 1")
    else:
        cur.execute("SELECT * FROM data_quality_reports WHERE id = %s", (report_id,))
    row = cur.fetchone()
    if not row:
        cur.close()
        conn.close()
        return None

    cur.execute("SELECT * FROM data_quality_column_results WHERE report_id = %s ORDER BY id", (row["id"],))
    profiles = cur.fetchall()
    cur.execute("SELECT COUNT(*) AS count FROM data_quality_issues WHERE report_id = %s", (row["id"],))
    issue_count = int(cur.fetchone()["count"])
    cur.close()
    conn.close()

    payload = dict(row)
    payload.update({
        "report_id": row["id"],
        "quality_level": get_quality_level(float(row["quality_score"])),
        "completeness_score": float(row["completeness_score"]),
        "uniqueness_score": float(row["uniqueness_score"]),
        "validity_score": float(row.get("validity_score") or 100),
        "quality_score": float(row["quality_score"]),
        "column_profiles": profiles,
        "issue_count": issue_count,
        "recommendations": build_recommendations(
            row["missing_values"], row["duplicate_rows"],
            float(row["completeness_score"]), float(row["uniqueness_score"]),
        ),
    })
    return payload


@app.post("/api/analyze-dataset")
async def analyze_dataset(
    request: Request,
    rules_json: str = Form("{}"),
    file: UploadFile = File(...),
):
    original_name, suffix, contents = await read_upload_limited(
        file,
        ALLOWED_DATASET_EXTENSIONS,
        max_upload_bytes=MAX_DATASET_UPLOAD_BYTES,
    )
    rules = parse_quality_rules(rules_json)
    try:
        if suffix == ".csv":
            df = pd.read_csv(io.BytesIO(contents))
        else:
            df = pd.read_excel(io.BytesIO(contents), engine="openpyxl")
    except Exception as exc:
        raise HTTPException(status_code=422, detail=f"Could not read dataset: {exc}") from exc

    total_rows = len(df)
    total_columns = len(df.columns)
    total_cells = total_rows * total_columns

    missing_values = int(df.isnull().sum().sum())
    duplicate_rows = int(df.duplicated().sum())

    completeness_score = 0 if total_cells == 0 else ((total_cells - missing_values) / total_cells) * 100
    uniqueness_score = 0 if total_rows == 0 else ((total_rows - duplicate_rows) / total_rows) * 100

    column_profiles, issues, validity_score = profile_dataframe(df, rules)
    quality_score = round((completeness_score * 0.45) + (uniqueness_score * 0.35) + (validity_score * 0.20), 2)
    quality_level = get_quality_level(quality_score)

    with engine.begin() as conn:
        report_id = conn.execute(text("""
            INSERT INTO data_quality_reports
            (asset_name, file_name, total_rows, total_columns, missing_values,
             duplicate_rows, completeness_score, uniqueness_score, validity_score, quality_score, applied_rules)
            VALUES
            (:asset_name, :file_name, :total_rows, :total_columns, :missing_values,
             :duplicate_rows, :completeness_score, :uniqueness_score, :validity_score, :quality_score, CAST(:applied_rules AS JSONB))
            RETURNING id
        """), {
            "asset_name": original_name,
            "file_name": original_name,
            "total_rows": total_rows,
            "total_columns": total_columns,
            "missing_values": missing_values,
            "duplicate_rows": duplicate_rows,
            "completeness_score": round(completeness_score, 2),
            "uniqueness_score": round(uniqueness_score, 2),
            "validity_score": validity_score,
            "quality_score": quality_score,
            "applied_rules": json.dumps(rules, ensure_ascii=False),
        }).scalar_one()

        if column_profiles:
            conn.execute(text("""
                INSERT INTO data_quality_column_results
                (report_id, column_name, data_type, total_values, missing_values, unique_values, duplicate_values, validity_score, pii_type)
                VALUES
                (:report_id, :column_name, :data_type, :total_values, :missing_values, :unique_values, :duplicate_values, :validity_score, :pii_type)
            """), [{**profile, "report_id": report_id} for profile in column_profiles])

        if issues:
            conn.execute(text("""
                INSERT INTO data_quality_issues
                (report_id, column_name, row_number, issue_type, issue_value)
                VALUES (:report_id, :column_name, :row_number, :issue_type, :issue_value)
            """), [{**issue, "report_id": report_id} for issue in issues])

    write_audit(
        request.state.current_user,
        "analyze_dataset",
        "data_quality_report",
        report_id,
        {"file_name": original_name, "rows": total_rows, "columns": total_columns, "quality_score": quality_score},
    )
    report = load_quality_report(report_id)
    report["quality_level"] = quality_level
    return report


@app.get("/api/data-quality-report")
def get_data_quality_report():
    report = load_quality_report()
    if report is None:
        raise HTTPException(status_code=404, detail="No data quality report is available")

    return report


@app.get("/api/data-quality-reports/{report_id}")
def get_data_quality_report_details(report_id: int):
    report = load_quality_report(report_id)
    if not report:
        raise HTTPException(status_code=404, detail="Data quality report not found")
    return report


@app.get("/api/data-quality-reports/{report_id}/issues.csv")
def export_data_quality_issues(report_id: int):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT column_name, row_number, issue_type, issue_value FROM data_quality_issues WHERE report_id = %s ORDER BY id", (report_id,))
    issues = cur.fetchall()
    cur.close()
    conn.close()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["column_name", "row_number", "issue_type", "issue_value"])
    for issue in issues:
        writer.writerow([issue["column_name"], issue["row_number"], issue["issue_type"], issue["issue_value"]])
    return Response(
        content="\ufeff" + output.getvalue(),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="data-quality-issues-{report_id}.csv"'},
    )

from fastapi import UploadFile, File, Form
from fastapi import HTTPException
import os
import json
import shutil
import csv
from urllib import request as urlrequest
from urllib.error import HTTPError, URLError
import psycopg2
from psycopg2.extras import RealDictCursor

EVIDENCE_UPLOAD_DIR = str(EVIDENCE_UPLOAD_PATH)
os.makedirs(EVIDENCE_UPLOAD_DIR, exist_ok=True)

AI_PROVIDER = os.getenv("AI_PROVIDER", "rule_based").lower()
AI_MODEL_NAME = os.getenv(
    "AI_MODEL_NAME",
    "qwen2.5:7b-instruct" if AI_PROVIDER == "ollama" else "ndmo-evidence-assessor-v1"
)
AI_BASE_MODEL_NAME = os.getenv(
    "AI_BASE_MODEL_NAME",
    "Qwen2.5 Instruct via Ollama" if AI_PROVIDER == "ollama" else "MoritzLaurer/mDeBERTa-v3-base-mnli-xnli"
)
OLLAMA_BASE_URL = os.getenv("OLLAMA_BASE_URL", "http://127.0.0.1:11434")
EMBEDDING_PROVIDER = os.getenv("EMBEDDING_PROVIDER", "ollama" if AI_PROVIDER == "ollama" else "disabled").lower()
EMBEDDING_MODEL_NAME = os.getenv("EMBEDDING_MODEL_NAME", "bge-m3")
EVIDENCE_RETRIEVAL_MODE = os.getenv(
    "EVIDENCE_RETRIEVAL_MODE",
    "semantic" if AI_PROVIDER == "ollama" else "lexical"
).lower()
SEMANTIC_RETRIEVAL_ENABLED = (
    AI_PROVIDER == "ollama"
    and EMBEDDING_PROVIDER == "ollama"
    and EVIDENCE_RETRIEVAL_MODE == "semantic"
)


def get_env_int(name, default):
    try:
        return int(os.getenv(name, default))
    except (TypeError, ValueError):
        return default


def get_env_float(name, default):
    try:
        return float(os.getenv(name, default))
    except (TypeError, ValueError):
        return default


MAX_AI_QUESTIONS = get_env_int("MAX_AI_QUESTIONS", 3)
OLLAMA_TIMEOUT_SECONDS = get_env_int("OLLAMA_TIMEOUT_SECONDS", 45)
OLLAMA_NUM_THREAD = get_env_int("OLLAMA_NUM_THREAD", 2)
OLLAMA_NUM_CTX = get_env_int("OLLAMA_NUM_CTX", 1536)
OLLAMA_NUM_PREDICT = get_env_int("OLLAMA_NUM_PREDICT", 120)
EMBEDDING_TIMEOUT_SECONDS = get_env_int("EMBEDDING_TIMEOUT_SECONDS", 120)
EMBEDDING_BATCH_SIZE = get_env_int("EMBEDDING_BATCH_SIZE", 32)
SEMANTIC_CACHE_FILES = get_env_int("SEMANTIC_CACHE_FILES", 3)
EVIDENCE_TOP_CHUNKS = get_env_int("EVIDENCE_TOP_CHUNKS", 3)
EVIDENCE_CANDIDATE_CHUNKS = get_env_int("EVIDENCE_CANDIDATE_CHUNKS", 8)
EVIDENCE_SNIPPET_CHARS = get_env_int("EVIDENCE_SNIPPET_CHARS", 280)
EVIDENCE_PROMPT_CHARS = get_env_int("EVIDENCE_PROMPT_CHARS", 1300)
EVIDENCE_MAX_TABLE_ROWS = get_env_int("EVIDENCE_MAX_TABLE_ROWS", 1200)
EVIDENCE_MAX_SHEET_ROWS = get_env_int("EVIDENCE_MAX_SHEET_ROWS", 400)
EVIDENCE_MIN_RELEVANCE_SCORE = get_env_float("EVIDENCE_MIN_RELEVANCE_SCORE", 0.12)
SEMANTIC_MIN_RELEVANCE_SCORE = get_env_float("SEMANTIC_MIN_RELEVANCE_SCORE", 0.0)

SEMANTIC_CHUNK_CACHE = {}


def create_ai_assessment_tables():
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS ndmo_document_chunks (
            id SERIAL PRIMARY KEY,
            evidence_id INTEGER REFERENCES ndmo_evidence(id) ON DELETE CASCADE,
            question_id INTEGER REFERENCES ndmo_questions(id) ON DELETE CASCADE,
            chunk_text TEXT NOT NULL,
            file_name TEXT,
            file_type TEXT,
            page_number INTEGER,
            sheet_name TEXT,
            row_number INTEGER,
            column_name TEXT,
            similarity_score NUMERIC(6, 4),
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS ndmo_ai_assessment_results (
            id SERIAL PRIMARY KEY,
            question_id INTEGER REFERENCES ndmo_questions(id) ON DELETE CASCADE,
            evidence_id INTEGER REFERENCES ndmo_evidence(id) ON DELETE CASCADE,
            ai_answer VARCHAR(20) NOT NULL,
            confidence INTEGER NOT NULL,
            reason TEXT,
            evidence_text TEXT,
            evidence_location TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS ndmo_ai_domain_assessment_runs (
            id SERIAL PRIMARY KEY,
            domain_id INTEGER REFERENCES ndmo_domains(id) ON DELETE CASCADE,
            evidence_id INTEGER REFERENCES ndmo_evidence(id) ON DELETE CASCADE,
            total_questions INTEGER NOT NULL,
            yes_count INTEGER NOT NULL DEFAULT 0,
            partial_count INTEGER NOT NULL DEFAULT 0,
            no_count INTEGER NOT NULL DEFAULT 0,
            domain_score NUMERIC(6, 2) NOT NULL DEFAULT 0,
            ndi_score NUMERIC(6, 2) NOT NULL DEFAULT 0,
            model_name TEXT NOT NULL DEFAULT 'ndmo-evidence-assessor-v1',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS ndmo_ai_training_examples (
            id SERIAL PRIMARY KEY,
            question_id INTEGER REFERENCES ndmo_questions(id) ON DELETE CASCADE,
            evidence_id INTEGER REFERENCES ndmo_evidence(id) ON DELETE CASCADE,
            evidence_text TEXT NOT NULL,
            evidence_location TEXT,
            predicted_answer VARCHAR(20),
            corrected_answer VARCHAR(20) NOT NULL,
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
    """)

    conn.commit()
    cur.close()
    conn.close()


@app.on_event("startup")
def startup_create_ai_assessment_tables():
    run_migrations()
    create_ai_assessment_tables()
    seed_reference_catalog()
    seed_demo_environment()


def get_question_by_id(question_id):
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            q.id,
            q.domain_id,
            q.question_code,
            q.question_text_ar,
            q.question_text_en,
            q.evidence_required,
            q.max_score,
            d.name_ar AS domain_name_ar,
            d.name_en AS domain_name_en
        FROM ndmo_questions q
        JOIN ndmo_domains d ON d.id = q.domain_id
        WHERE q.id = %s;
    """, (question_id,))

    question = cur.fetchone()
    cur.close()
    conn.close()
    return question


def get_domain_by_id(domain_id):
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT id, name_ar, name_en, description, order_number
        FROM ndmo_domains
        WHERE id = %s;
    """, (domain_id,))

    domain = cur.fetchone()
    cur.close()
    conn.close()
    return domain


def get_questions_by_domain_id(domain_id):
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            q.id,
            q.domain_id,
            q.question_code,
            q.question_text_ar,
            q.question_text_en,
            q.evidence_required,
            q.max_score,
            d.name_ar AS domain_name_ar,
            d.name_en AS domain_name_en
        FROM ndmo_questions q
        JOIN ndmo_domains d ON d.id = q.domain_id
        WHERE q.domain_id = %s
        ORDER BY q.id;
    """, (domain_id,))

    questions = cur.fetchall()
    cur.close()
    conn.close()
    return questions


def get_all_ndmo_questions():
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            q.id,
            q.domain_id,
            q.question_code,
            q.question_text_ar,
            q.question_text_en,
            q.evidence_required,
            q.max_score,
            d.name_ar AS domain_name_ar,
            d.name_en AS domain_name_en
        FROM ndmo_questions q
        JOIN ndmo_domains d ON d.id = q.domain_id
        ORDER BY q.domain_id, q.id;
    """)

    questions = cur.fetchall()
    cur.close()
    conn.close()
    return questions


def get_evidence_by_id(evidence_id):
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT id, question_id, file_name, file_path, evidence_type, status, uploaded_at
        FROM ndmo_evidence
        WHERE id = %s;
    """, (evidence_id,))

    evidence = cur.fetchone()
    cur.close()
    conn.close()
    return evidence


def clean_text(value):
    if value is None:
        return ""
    return str(value).strip()


def normalize_training_text(value):
    return re.sub(r"\s+", " ", clean_text(value))


def normalize_evidence_text(value):
    return re.sub(r"\s+", " ", clean_text(value))


def truncate_text(value, max_chars):
    text_value = normalize_evidence_text(value)
    if len(text_value) <= max_chars:
        return text_value

    truncated = text_value[:max_chars].rsplit(" ", 1)[0]
    return f"{truncated}..."


def normalize_label(value):
    return normalize_domain_value(value).replace(" ", "_")


IMPORTANT_EVIDENCE_LABELS = [
    "ndmo_domain",
    "domain",
    "business_domain",
    "data_domain",
    "evidence_type",
    "approval_status",
    "evidence_status",
    "assessment_status",
    "model_decision",
    "data_owner",
    "document_owner",
    "policy_owner",
    "data_steward",
    "quality_score",
    "completeness_pct",
    "uniqueness_pct",
    "accuracy_pct",
    "review_date",
    "evidence_date",
    "requirement_reference",
    "evidence_id",
    "policy_id",
    "policy_name",
    "مجال_ndmo",
    "مجال_الاعمال",
    "مجال_البيانات",
    "المجال",
    "نوع_الدليل",
    "نوع_الوثيقة",
    "نوع_المستند",
    "نوع_السياسة",
    "حالة_الاعتماد",
    "حالة_الموافقة",
    "حالة_الدليل",
    "مالك_البيانات",
    "مالك_الدليل",
    "مالك_المستند",
    "مالك_السياسة",
    "امين_البيانات",
    "المسؤول",
    "درجة_الجودة",
    "نسبة_الاكتمال",
    "نسبة_الدقة",
    "نسبة_التفرد",
    "تاريخ_المراجعة",
    "مؤشر_الاداء",
    "خطة_التحسين",
]


def prioritize_row_values(row_values):
    def priority(row_value):
        label = normalize_label(row_value.split(":", 1)[0])
        for index, important_label in enumerate(IMPORTANT_EVIDENCE_LABELS):
            if label == important_label or important_label in label:
                return index
        return len(IMPORTANT_EVIDENCE_LABELS) + 1

    return sorted(row_values, key=priority)


def build_stratified_row_sample(first_row, last_row, limit):
    """Return a bounded sample that covers the beginning and the full sheet."""
    try:
        first_row = int(first_row)
        last_row = int(last_row)
        limit = int(limit)
    except (TypeError, ValueError):
        return None

    total_rows = max(0, last_row - first_row + 1)
    if limit <= 0:
        return set()
    if total_rows <= limit:
        return None

    leading_count = min(total_rows, max(1, limit // 4))
    selected_rows = set(range(first_row, first_row + leading_count))
    remaining_count = limit - len(selected_rows)
    sample_start = first_row + leading_count

    if remaining_count <= 0 or sample_start > last_row:
        return selected_rows

    if remaining_count == 1:
        selected_rows.add(last_row)
    else:
        step = (last_row - sample_start) / (remaining_count - 1)
        for index in range(remaining_count):
            selected_rows.add(round(sample_start + (index * step)))

    if len(selected_rows) < limit:
        for row_number in range(sample_start, last_row + 1):
            selected_rows.add(row_number)
            if len(selected_rows) >= limit:
                break

    return selected_rows


def column_index_to_letter(index):
    letters = ""
    while index >= 0:
        index, remainder = divmod(index, 26)
        letters = chr(65 + remainder) + letters
        index -= 1
    return letters


def get_question_text(question):
    return clean_text(question.get("question_text_en")) or clean_text(question.get("question_text_ar"))


def normalize_domain_value(value):
    text_value = normalize_evidence_text(value).lower()
    text_value = text_value.translate(str.maketrans({
        "أ": "ا",
        "إ": "ا",
        "آ": "ا",
        "ى": "ي",
    }))
    text_value = re.sub(r"[\u064B-\u065F\u0670]", "", text_value)
    text_value = text_value.replace("_", " ").replace("-", " ")
    return re.sub(r"\s+", " ", text_value).strip()


DOMAIN_ALIASES = {
    "data governance": {"data governance", "governance"},
    "data catalog and metadata": {"data catalog", "metadata", "catalog"},
    "data quality": {"data quality", "quality"},
    "data classification": {"data classification", "classification"},
    "data privacy": {"data privacy", "privacy"},
    "data sharing": {"data sharing", "sharing"},
    "master and reference data": {"master data", "reference data", "master reference"},
    "business intelligence and analytics": {"business intelligence", "analytics", "bi"},
    "data architecture": {"data architecture", "architecture"},
    "data operations": {"data operations", "operations"},
    "records management": {"records management", "records"},
    "open data": {"open data"},
    "freedom of information": {"freedom of information", "information requests"},
    "data value realization": {"data value", "value realization"},
    "حوكمة البيانات": {"حوكمة البيانات", "حوكمة"},
    "كتالوج البيانات والبيانات الوصفية": {"كتالوج البيانات", "البيانات الوصفية", "الكتالوج"},
    "جودة البيانات": {"جودة البيانات", "الجودة"},
    "تصنيف البيانات": {"تصنيف البيانات", "التصنيف"},
    "خصوصية البيانات": {"خصوصية البيانات", "الخصوصية"},
    "مشاركة البيانات": {"مشاركة البيانات", "المشاركة"},
    "البيانات الرئيسية والمرجعية": {"البيانات الرئيسية", "البيانات المرجعية"},
    "ذكاء الاعمال والتحليلات": {"ذكاء الاعمال", "التحليلات"},
    "معمارية البيانات": {"معمارية البيانات", "المعمارية"},
    "عمليات البيانات": {"عمليات البيانات", "العمليات"},
    "ادارة السجلات": {"ادارة السجلات", "السجلات"},
    "البيانات المفتوحة": {"البيانات المفتوحة"},
    "حرية المعلومات": {"حرية المعلومات", "طلبات المعلومات"},
    "تحقيق قيمة البيانات": {"تحقيق قيمة البيانات", "قيمة البيانات"},
}


def build_question_domain_context(question):
    domain_names = [
        clean_text(question.get("domain_name_en")),
        clean_text(question.get("domain_name_ar")),
    ]
    phrases = set()

    for domain_name in domain_names:
        normalized_name = normalize_domain_value(domain_name)
        if normalized_name:
            phrases.add(normalized_name)
            phrases.update(DOMAIN_ALIASES.get(normalized_name, set()))

    tokens = set()
    for phrase in phrases:
        tokens.update(
            token for token in tokenize(phrase)
            if token not in {"data", "domain", "management", "البيانات", "بيانات", "مجال", "ادارة", "إدارة"}
        )

    return {
        "names": [name for name in domain_names if name],
        "phrases": phrases,
        "tokens": tokens,
    }


def get_domain_label_values_from_text(text_value):
    normalized_text = normalize_evidence_text(text_value)
    label_patterns = [
        r"(?:^|[;\n])\s*(?:NDMO[_\s-]*Domain|Business[_\s-]*Domain|Data[_\s-]*Domain|Domain)\s*:\s*([^;\n]+)",
        r"(?:^|[؛;\n])\s*(?:مجال\s*NDMO|مجال\s*الأعمال|مجال\s*الاعمال|مجال\s*البيانات|نطاق\s*البيانات|النطاق|المجال|مجال)\s*:\s*([^؛;\n]+)",
    ]
    values = []

    for pattern in label_patterns:
        values.extend(
            clean_text(match)
            for match in re.findall(pattern, normalized_text, flags=re.IGNORECASE)
            if clean_text(match)
        )

    return values


def text_matches_domain(value, domain_context):
    normalized_value = normalize_domain_value(value)
    if not normalized_value:
        return False

    for phrase in domain_context.get("phrases", set()):
        normalized_phrase = normalize_domain_value(phrase)
        if normalized_phrase and (
            normalized_phrase in normalized_value
            or normalized_value in normalized_phrase
        ):
            return True

    value_tokens = set(tokenize(normalized_value))
    domain_tokens = domain_context.get("tokens", set())
    return bool(value_tokens & domain_tokens)


def get_chunk_domain_alignment(chunk, domain_context):
    if not domain_context.get("phrases") and not domain_context.get("tokens"):
        return "unknown"

    domain_values = get_domain_label_values_from_text(chunk.get("text", ""))
    column_name = normalize_domain_value(chunk.get("column_name"))

    if column_name in {
        "domain", "ndmo domain", "business domain", "data domain",
        "المجال", "مجال", "مجال ndmo", "مجال الاعمال", "مجال الأعمال",
        "مجال البيانات", "نطاق البيانات", "النطاق"
    }:
        cell_text = clean_text(chunk.get("text"))
        if ":" in cell_text:
            domain_values.append(cell_text.split(":", 1)[1])
        else:
            domain_values.append(cell_text)

    if domain_values:
        return "match" if any(
            text_matches_domain(value, domain_context)
            for value in domain_values
        ) else "mismatch"

    search_text = " ".join([
        clean_text(chunk.get("sheet_name")),
        clean_text(chunk.get("column_name")),
        clean_text(chunk.get("text")),
    ])

    if text_matches_domain(search_text, domain_context):
        return "match"

    return "unknown"


def extract_text_from_excel(file_path):
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise HTTPException(
            status_code=500,
            detail="Excel analysis requires openpyxl. Install it with: pip install openpyxl"
        ) from exc

    workbook = load_workbook(file_path, data_only=True, read_only=True)
    items = []

    for sheet in workbook.worksheets:
        headers = []
        header_row_number = None
        sampled_data_rows = None
        processed_data_rows = 0

        for row in sheet.iter_rows():
            cells = list(row)
            cell_values = [clean_text(cell.value) for cell in cells]
            non_empty_values = [value for value in cell_values if value]
            row_number = next(
                (
                    getattr(cell, "row", None)
                    for cell in cells
                    if getattr(cell, "row", None) is not None
                ),
                None
            )

            if not non_empty_values or row_number is None:
                continue

            if (
                header_row_number is not None
                and row_number != header_row_number
                and sampled_data_rows is not None
                and row_number not in sampled_data_rows
            ):
                continue

            if header_row_number is not None and row_number != header_row_number:
                processed_data_rows += 1
                if processed_data_rows > EVIDENCE_MAX_SHEET_ROWS:
                    break

            if headers == [] and len(non_empty_values) >= 2:
                header_row_number = row_number
                headers = [
                    value or get_column_letter(index)
                    for index, value in enumerate(cell_values, start=1)
                ]
                sampled_data_rows = build_stratified_row_sample(
                    header_row_number + 1,
                    sheet.max_row,
                    EVIDENCE_MAX_SHEET_ROWS
                )

            row_values = []
            used_columns = []

            for index, cell in enumerate(cells, start=1):
                value = clean_text(cell.value)
                if not value:
                    continue

                column_name = get_column_letter(index)
                header_name = ""
                if headers and index - 1 < len(headers) and row_number != header_row_number:
                    header_name = clean_text(headers[index - 1])

                label = header_name or column_name
                row_values.append(f"{label}: {value}")
                used_columns.append(label)
                items.append({
                    "text": f"{label}: {value}",
                    "file_name": os.path.basename(file_path),
                    "file_type": Path(file_path).suffix.lower(),
                    "page_number": None,
                    "sheet_name": sheet.title,
                    "row_number": row_number,
                    "column_name": label,
                    "chunk_level": "header" if row_number == header_row_number else "cell",
                    "non_empty_count": 1,
                    "location": f"Sheet: {sheet.title}, Row: {row_number}, Column: {label}"
                })

            if row_values:
                row_level = "header" if row_number == header_row_number else "row"
                prioritized_row_values = prioritize_row_values(row_values)
                items.append({
                    "text": truncate_text("; ".join(prioritized_row_values), 900),
                    "search_text": truncate_text("; ".join(prioritized_row_values), 2200),
                    "file_name": os.path.basename(file_path),
                    "file_type": Path(file_path).suffix.lower(),
                    "page_number": None,
                    "sheet_name": sheet.title,
                    "row_number": row_number,
                    "column_name": ", ".join(used_columns[:8]),
                    "chunk_level": row_level,
                    "non_empty_count": len(row_values),
                    "location": f"Sheet: {sheet.title}, Row: {row_number}, Columns: {', '.join(used_columns[:8])}"
                })

    workbook.close()
    return items


def extract_text_from_csv(file_path):
    items = []

    with open(file_path, "r", encoding="utf-8-sig", errors="ignore", newline="") as file:
        sample = file.read(4096)
        file.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample)
        except csv.Error:
            dialect = csv.excel

        reader = csv.reader(file, dialect)
        rows = []
        for row_index, row in enumerate(reader, start=1):
            rows.append(row)
            if row_index >= EVIDENCE_MAX_TABLE_ROWS + 1:
                break

    if not rows:
        return items

    headers = [
        clean_text(value) or column_index_to_letter(index)
        for index, value in enumerate(rows[0])
    ]
    file_name = os.path.basename(file_path)
    file_type = Path(file_path).suffix.lower()

    header_text = "; ".join(
        f"{column_index_to_letter(index)}: {header}"
        for index, header in enumerate(headers)
        if header
    )
    if header_text:
        items.append({
            "text": truncate_text(header_text, 900),
            "search_text": truncate_text(header_text, 2200),
            "file_name": file_name,
            "file_type": file_type,
            "page_number": None,
            "sheet_name": "CSV",
            "row_number": 1,
            "column_name": ", ".join(headers[:8]),
            "chunk_level": "header",
            "non_empty_count": len([header for header in headers if header]),
            "location": f"CSV, Row: 1, Columns: {', '.join(headers[:8])}"
        })

    for row_number, row in enumerate(rows[1:], start=2):
        row_values = []
        used_columns = []

        for index, value in enumerate(row):
            cleaned_value = clean_text(value)
            if not cleaned_value:
                continue

            header_name = headers[index] if index < len(headers) else column_index_to_letter(index)
            row_values.append(f"{header_name}: {cleaned_value}")
            used_columns.append(header_name)
            items.append({
                "text": f"{header_name}: {cleaned_value}",
                "file_name": file_name,
                "file_type": file_type,
                "page_number": None,
                "sheet_name": "CSV",
                "row_number": row_number,
                "column_name": header_name,
                "chunk_level": "cell",
                "non_empty_count": 1,
                "location": f"CSV, Row: {row_number}, Column: {header_name}"
            })

        if row_values:
            prioritized_row_values = prioritize_row_values(row_values)
            items.append({
                "text": truncate_text("; ".join(prioritized_row_values), 900),
                "search_text": truncate_text("; ".join(prioritized_row_values), 2200),
                "file_name": file_name,
                "file_type": file_type,
                "page_number": None,
                "sheet_name": "CSV",
                "row_number": row_number,
                "column_name": ", ".join(used_columns[:8]),
                "chunk_level": "row",
                "non_empty_count": len(row_values),
                "location": f"CSV, Row: {row_number}, Columns: {', '.join(used_columns[:8])}"
            })

    return items


def extract_text_from_pdf(file_path):
    items = []

    try:
        import fitz

        document = fitz.open(file_path)
        for page_index, page in enumerate(document, start=1):
            text_value = clean_text(page.get_text())
            if text_value:
                items.append({
                    "text": text_value,
                    "file_name": os.path.basename(file_path),
                    "file_type": Path(file_path).suffix.lower(),
                    "page_number": page_index,
                    "sheet_name": None,
                    "row_number": None,
                    "column_name": None,
                    "location": f"Page: {page_index}"
                })
        document.close()
        return items
    except ImportError:
        pass

    try:
        from pypdf import PdfReader

        reader = PdfReader(file_path)
        for page_index, page in enumerate(reader.pages, start=1):
            text_value = clean_text(page.extract_text())
            if text_value:
                items.append({
                    "text": text_value,
                    "file_name": os.path.basename(file_path),
                    "file_type": Path(file_path).suffix.lower(),
                    "page_number": page_index,
                    "sheet_name": None,
                    "row_number": None,
                    "column_name": None,
                    "location": f"Page: {page_index}"
                })
        return items
    except ImportError as exc:
        raise HTTPException(
            status_code=500,
            detail="PDF analysis requires PyMuPDF or pypdf. Install one with: pip install pypdf"
        ) from exc


def extract_text_from_plain_file(file_path):
    with open(file_path, "r", encoding="utf-8", errors="ignore") as file:
        text_value = clean_text(file.read())

    if not text_value:
        return []

    return [{
        "text": text_value,
        "file_name": os.path.basename(file_path),
        "file_type": Path(file_path).suffix.lower(),
        "page_number": None,
        "sheet_name": None,
        "row_number": None,
        "column_name": None,
        "location": "Text file"
    }]


def chunk_extracted_text(items, max_words=90):
    chunks = []

    for item in items:
        words = clean_text(item.get("text")).split()
        if not words:
            continue

        if len(words) <= max_words:
            chunks.append(item)
            continue

        for start in range(0, len(words), max_words):
            chunk_words = words[start:start + max_words]
            chunk = item.copy()
            chunk["text"] = " ".join(chunk_words)
            chunks.append(chunk)

    return chunks


def tokenize(text_value):
    stop_words = {
        "a", "an", "and", "are", "as", "at", "be", "by", "for", "from",
        "has", "have", "in", "is", "it", "of", "on", "or", "that", "the",
        "this", "to", "with", "there", "does", "domain", "requirement",
        "question", "periodically", "هل", "في", "من", "على", "عن", "الى",
        "إلى", "هذا", "هذه", "ذلك", "توجد", "يوجد"
    }
    normalized_text = clean_text(text_value).lower().translate(str.maketrans({
        "أ": "ا",
        "إ": "ا",
        "آ": "ا",
        "ى": "ي",
    }))
    normalized_text = re.sub(r"[\u064B-\u065F\u0670]", "", normalized_text)
    words = re.findall(r"[\w\u0600-\u06FF]+", normalized_text)
    return [word for word in words if len(word) > 2 and word not in stop_words]


def cosine_similarity(text_a, text_b):
    tokens_a = tokenize(text_a)
    tokens_b = tokenize(text_b)

    if not tokens_a or not tokens_b:
        return 0

    counts_a = Counter(tokens_a)
    counts_b = Counter(tokens_b)
    common_tokens = set(counts_a) & set(counts_b)
    dot_product = sum(counts_a[token] * counts_b[token] for token in common_tokens)
    magnitude_a = math.sqrt(sum(value * value for value in counts_a.values()))
    magnitude_b = math.sqrt(sum(value * value for value in counts_b.values()))

    if magnitude_a == 0 or magnitude_b == 0:
        return 0

    return dot_product / (magnitude_a * magnitude_b)


def expand_question_intent_terms(question_text):
    text_value = normalize_evidence_text(question_text).lower()
    terms = set(tokenize(text_value))

    if any(term in text_value for term in ["policy", "procedure", "documented", "سياسة", "إجراء", "موثق"]):
        terms.update({
            "policy", "procedure", "standard", "approved", "documented",
            "governance", "control", "سياسة", "إجراء", "معتمد", "موثق"
        })

    if any(term in text_value for term in ["owner", "responsible", "steward", "مالك", "مسؤول"]):
        terms.update({
            "owner", "steward", "responsible", "accountable", "custodian",
            "مالك", "مسؤول", "أمين", "صاحب"
        })

    if any(term in text_value for term in ["measured", "improved", "periodic", "kpi", "قياس", "تحسين", "دوري"]):
        terms.update({
            "measure", "measured", "metric", "kpi", "improvement", "improved",
            "review", "periodic", "quarterly", "قياس", "مؤشر", "تحسين", "دوري"
        })

    return terms


def get_chunk_search_text(chunk):
    return " ".join([
        clean_text(chunk.get("sheet_name")),
        clean_text(chunk.get("column_name")),
        clean_text(chunk.get("location")),
        clean_text(chunk.get("search_text")),
        clean_text(chunk.get("text"))
    ])


def get_labeled_values(text_value, labels):
    normalized_text = normalize_evidence_text(text_value)
    values = []

    for label in labels:
        label_pattern = re.escape(label).replace("_", r"[_\s-]*")
        pattern = rf"(?:^|[؛;\n])\s*{label_pattern}\s*:\s*([^؛;\n]+)"
        values.extend(
            clean_text(match)
            for match in re.findall(pattern, normalized_text, flags=re.IGNORECASE)
            if clean_text(match)
        )

    return values


def any_normalized_value_contains(values, accepted_terms):
    normalized_terms = {normalize_domain_value(term) for term in accepted_terms}
    return any(
        accepted_term in normalize_domain_value(value)
        for value in values
        for accepted_term in normalized_terms
        if accepted_term
    )


EMPTY_EVIDENCE_VALUES = {
    "", "-", "--", "n/a", "na", "none", "null", "unknown", "not assigned",
    "not available", "غير متوفر", "غير متاح", "غير محدد", "غير معين", "لا يوجد"
}

POLICY_DOCUMENT_LABELS = [
    "Evidence_Type", "Evidence Type", "Policy Type", "Document Type",
    "Policy_Name", "Policy Name", "Policy_ID", "Policy ID", "Procedure_Name",
    "نوع الدليل", "نوع الوثيقة", "نوع المستند", "نوع السياسة",
    "اسم السياسة", "رقم السياسة", "اسم الإجراء", "اسم الاجراء"
]

APPROVAL_STATUS_LABELS = [
    "Approval_Status", "Approval Status", "Evidence_Status", "Evidence Status",
    "Assessment_Status", "Assessment Status", "حالة الاعتماد", "حالة الموافقة",
    "حالة الدليل", "حالة التقييم"
]

OWNER_LABELS = [
    "Data_Owner", "Data Owner", "Document_Owner", "Document Owner",
    "Policy_Owner", "Policy Owner", "Owner", "Data_Steward", "Data Steward",
    "Data_Custodian", "Data Custodian", "Responsible", "Accountable",
    "مالك البيانات", "مالك الدليل", "مالك المستند", "مالك السياسة",
    "أمين البيانات", "امين البيانات", "حافظ البيانات", "المسؤول", "المالك"
]

MEASUREMENT_LABELS = [
    "Quality_Score", "Quality Score", "Completeness_Pct", "Accuracy_Pct",
    "Uniqueness_Pct", "Review_Date", "Review Date", "Assessment_Date",
    "Improvement_Plan", "KPI", "Metric", "Measurement", "درجة الجودة",
    "نسبة الاكتمال", "نسبة الدقة", "نسبة التفرد", "تاريخ المراجعة",
    "تاريخ التقييم", "خطة التحسين", "مؤشر الأداء", "مؤشر الاداء", "القياس"
]

DOMAIN_IMPLEMENTATION_LABELS = [
    (
        {"metadata management", "data catalog and metadata", "ادارة البيانات الوصفية", "كتالوج البيانات والبيانات الوصفية"},
        ["Entity_Name", "Attribute_ID", "Attribute_Name", "Technical_Field_Name", "Attribute_Description_EN", "Attribute_Description_AR"]
    ),
    (
        {"data quality", "جودة البيانات"},
        ["Is_Mandatory", "Is_Primary_Key", "Is_Foreign_Key", "Is_Critical_Data_Attribute", "Quality_Score"]
    ),
    (
        {"data classification", "تصنيف البيانات"},
        ["Data_Classification_Level", "Classification_Criteria", "مستوى تصنيف البيانات", "معايير التصنيف"]
    ),
    (
        {"personal data protection", "data privacy", "حماية البيانات الشخصية", "خصوصية البيانات"},
        ["Is_PII", "Contains_Personal_Data", "يحتوي بيانات شخصية", "بيانات شخصية"]
    ),
    (
        {"data architecture", "معمارية البيانات"},
        ["System_Name", "Schema_Name", "Table_Name", "System_Type", "اسم النظام", "اسم المخطط", "اسم الجدول"]
    ),
    (
        {"reference and master data", "master and reference data", "البيانات المرجعية والرئيسية", "البيانات الرئيسية والمرجعية"},
        ["Master_or_Reference_Data", "Is_Master_Data", "Is_Reference_Data", "البيانات الرئيسية أو المرجعية"]
    ),
    (
        {"data integration", "تكامل البيانات"},
        ["Data_Source_Description", "Source_System", "Target_System", "Integration_Name", "وصف مصدر البيانات"]
    ),
]


def is_meaningful_evidence_value(value, allow_boolean=False):
    normalized_value = normalize_domain_value(value)
    if normalized_value in EMPTY_EVIDENCE_VALUES:
        return False
    if not allow_boolean and normalized_value in {"no", "false", "لا", "كلا"}:
        return False
    return bool(normalized_value)


def get_values_from_chunks(chunks, labels):
    values = []
    seen = set()

    for chunk in chunks:
        for text_key in ["text", "search_text"]:
            for value in get_labeled_values(chunk.get(text_key, ""), labels):
                normalized_value = normalize_domain_value(value)
                if normalized_value in seen:
                    continue
                seen.add(normalized_value)
                values.append(value)

    return values


def detect_question_intent(question_text):
    normalized_question = normalize_domain_value(question_text)

    if any(term in normalized_question for term in [
        "policy", "procedure", "documented", "سياسة", "اجراء", "موثق"
    ]):
        return "policy"
    if any(term in normalized_question for term in [
        "owner", "responsible", "steward", "assigned", "مالك", "مسؤول", "امين"
    ]):
        return "owner"
    if any(term in normalized_question for term in [
        "measured", "improved", "periodic", "kpi", "قياس", "تحسين", "دوري"
    ]):
        return "measurement"

    return "general"


def text_contains_any_term(text_value, terms):
    normalized_text = normalize_domain_value(text_value)
    return any(normalize_domain_value(term) in normalized_text for term in terms)


def get_domain_implementation_labels(domain_context):
    normalized_names = {
        normalize_domain_value(name)
        for name in domain_context.get("names", [])
        if clean_text(name)
    }
    labels = []

    for domain_names, domain_labels in DOMAIN_IMPLEMENTATION_LABELS:
        normalized_domain_names = {normalize_domain_value(name) for name in domain_names}
        if any(
            expected_name in current_name or current_name in expected_name
            for current_name in normalized_names
            for expected_name in normalized_domain_names
        ):
            labels.extend(domain_labels)

    return labels


def has_domain_implementation_signal(chunks, domain_context):
    labels = get_domain_implementation_labels(domain_context)
    if not labels:
        return False

    values = get_values_from_chunks(chunks, labels)
    return any(is_meaningful_evidence_value(value, allow_boolean=True) for value in values)


def score_domain_implementation_evidence(chunk, domain_context):
    labels = get_domain_implementation_labels(domain_context)
    if not labels or chunk.get("chunk_level") == "header":
        return 0

    values = get_values_from_chunks([chunk], labels)
    if any(is_meaningful_evidence_value(value, allow_boolean=True) for value in values):
        return 0.18

    return 0


def score_requirement_specific_evidence(question_text, chunk):
    text_value = get_chunk_search_text(chunk)
    normalized_question = normalize_evidence_text(question_text).lower()
    normalized_text = normalize_evidence_text(text_value).lower()
    score = 0

    if any(term in normalized_question for term in ["policy", "procedure", "documented", "سياسة", "اجراء", "إجراء", "موثق"]):
        evidence_types = get_labeled_values(text_value, [
            "Evidence_Type", "Evidence Type", "Policy Type", "Document Type",
            "نوع الدليل", "نوع الوثيقة", "نوع المستند", "نوع السياسة"
        ])
        approval_values = get_labeled_values(text_value, [
            "Approval_Status", "Approval Status", "Evidence_Status", "Evidence Status",
            "Assessment_Status", "Assessment Status", "حالة الاعتماد", "حالة الموافقة",
            "حالة الدليل", "حالة التقييم"
        ])

        if any_normalized_value_contains(evidence_types, {"policy", "procedure", "standard", "سياسة", "اجراء", "إجراء", "معيار"}):
            score += 0.32
        elif any(term in normalized_text for term in ["policy", "procedure", "standard", "سياسة", "اجراء", "إجراء", "معيار"]):
            score += 0.14

        if any_normalized_value_contains(approval_values, {"approved", "accepted", "valid", "active", "معتمد", "معتمدة", "مقبول", "مقبولة", "ساري", "سارية"}):
            score += 0.12

    if any(term in normalized_question for term in ["owner", "responsible", "steward", "مالك", "مسؤول"]):
        owner_values = get_labeled_values(text_value, [
            "Data_Owner", "Data Owner", "Document_Owner", "Document Owner",
            "Policy_Owner", "Policy Owner", "Owner", "Data_Steward", "Data Steward",
            "Responsible", "Accountable", "مالك البيانات", "مالك الدليل", "مالك المستند",
            "مالك السياسة", "أمين البيانات", "امين البيانات", "المسؤول", "المالك"
        ])
        if any(clean_text(value) for value in owner_values):
            score += 0.34
        elif any(term in normalized_text for term in ["owner", "steward", "responsible", "مالك", "مسؤول"]):
            score += 0.14

    if any(term in normalized_question for term in ["measured", "improved", "periodic", "kpi", "قياس", "تحسين", "دوري"]):
        measurement_values = get_labeled_values(text_value, [
            "Quality_Score", "Quality Score", "Completeness_Pct", "Accuracy_Pct",
            "Uniqueness_Pct", "Review_Date", "Review Date", "Model_Decision",
            "Assessment_Status", "Improvement_Plan", "KPI", "Metric", "Measurement",
            "درجة الجودة", "نسبة الاكتمال", "نسبة الدقة", "نسبة التفرد",
            "تاريخ المراجعة", "قرار الموديل", "حالة التقييم", "خطة التحسين",
            "مؤشر الأداء", "مؤشر الاداء", "القياس"
        ])
        if any(clean_text(value) for value in measurement_values):
            score += 0.3
        elif any(term in normalized_text for term in [
            "quality_score", "quality score", "kpi", "metric", "review", "improvement",
            "درجة الجودة", "مؤشر", "مراجعة", "تحسين", "قياس", "اكتمال", "دقة"
        ]):
            score += 0.13

    return score


def score_chunk_for_question(question_text, chunk, domain_context=None):
    domain_context = domain_context or {"phrases": set(), "tokens": set()}
    domain_alignment = get_chunk_domain_alignment(chunk, domain_context)
    if domain_alignment == "mismatch":
        return 0

    search_text = get_chunk_search_text(chunk)
    base_score = cosine_similarity(question_text, search_text)
    chunk_tokens = set(tokenize(search_text))
    intent_terms = expand_question_intent_terms(question_text)
    intent_overlap = len(chunk_tokens & intent_terms)
    score = base_score + min(0.28, intent_overlap * 0.055)
    score += score_requirement_specific_evidence(question_text, chunk)
    score += score_domain_implementation_evidence(chunk, domain_context)

    if domain_alignment == "match":
        score += 0.22

    chunk_level = chunk.get("chunk_level")
    if chunk_level == "row":
        score += 0.025
    elif chunk_level == "header":
        score = min(score * 0.25, 0.08)

    text_length = len(normalize_evidence_text(chunk.get("text")))
    non_empty_count = int(chunk.get("non_empty_count") or 0)
    if text_length > 700:
        score -= 0.08
    elif text_length > 420:
        score -= 0.035

    if non_empty_count > 18:
        score -= 0.055
    elif non_empty_count > 10:
        score -= 0.025

    return max(0, round(score, 4))


def choose_best_sheet(scored_chunks):
    sheet_scores = {}

    for chunk in scored_chunks:
        sheet_name = chunk.get("sheet_name")
        if not sheet_name:
            continue

        sheet_scores.setdefault(sheet_name, []).append(chunk.get("similarity_score", 0))

    if not sheet_scores:
        return None

    ranked_sheets = []
    for sheet_name, scores in sheet_scores.items():
        top_scores = sorted(scores, reverse=True)[:5]
        sheet_score = max(top_scores) + (sum(top_scores) / max(len(top_scores), 1)) * 0.35
        ranked_sheets.append((sheet_score, sheet_name))

    ranked_sheets.sort(reverse=True)
    if ranked_sheets[0][0] <= 0:
        return None

    return ranked_sheets[0][1]


def get_chunk_group_key(chunk):
    if chunk.get("sheet_name") and chunk.get("row_number") is not None:
        return ("sheet-row", chunk.get("sheet_name"), chunk.get("row_number"))
    if chunk.get("page_number") is not None:
        return ("page", chunk.get("page_number"))
    return ("text", truncate_text(chunk.get("text"), 80))


def find_relevant_evidence(question_text, extracted_chunks, domain_context=None):
    domain_context = domain_context or {"phrases": set(), "tokens": set()}
    scored_chunks = []

    for chunk in extracted_chunks:
        scored_chunk = chunk.copy()
        if "similarity_score" not in scored_chunk:
            scored_chunk["similarity_score"] = score_chunk_for_question(
                question_text,
                scored_chunk,
                domain_context=domain_context
            )
        scored_chunks.append(scored_chunk)

    semantic_retrieval = any(
        chunk.get("retrieval_method") == "semantic"
        for chunk in scored_chunks
    )
    if not semantic_retrieval:
        domain_matched_chunks = [
            chunk for chunk in scored_chunks
            if get_chunk_domain_alignment(chunk, domain_context) == "match"
        ]
        if domain_matched_chunks:
            scored_chunks = domain_matched_chunks

        best_sheet = choose_best_sheet(scored_chunks)
        if best_sheet:
            scored_chunks = [
                chunk for chunk in scored_chunks
                if chunk.get("sheet_name") in [best_sheet, None, ""]
            ]

    best_chunks_by_group = {}
    row_chunks_by_group = {}
    minimum_score = (
        SEMANTIC_MIN_RELEVANCE_SCORE
        if semantic_retrieval
        else EVIDENCE_MIN_RELEVANCE_SCORE
    )
    for chunk in scored_chunks:
        if chunk.get("similarity_score", 0) < minimum_score:
            continue

        group_key = get_chunk_group_key(chunk)
        if chunk.get("chunk_level") in ["row", "header"]:
            row_chunks_by_group[group_key] = chunk

        current_chunk = best_chunks_by_group.get(group_key)
        if current_chunk is None:
            best_chunks_by_group[group_key] = chunk
            continue

        chunk_score = chunk.get("similarity_score", 0)
        current_score = current_chunk.get("similarity_score", 0)
        chunk_is_row = chunk.get("chunk_level") in ["row", "header"]
        current_is_row = current_chunk.get("chunk_level") in ["row", "header"]

        if chunk_score > current_score or (
            chunk_is_row and not current_is_row and chunk_score >= current_score - 0.08
        ):
            best_chunks_by_group[group_key] = chunk

    ranked_chunks = sorted(
        best_chunks_by_group.values(),
        key=lambda item: (
            item.get("similarity_score", 0),
            1 if item.get("chunk_level") in ["row", "header"] else 0
        ),
        reverse=True
    )

    relevant_chunks = []
    for chunk in ranked_chunks:
        group_key = get_chunk_group_key(chunk)
        row_chunk = row_chunks_by_group.get(group_key)
        if (
            chunk.get("chunk_level") == "cell"
            and row_chunk is not None
            and row_chunk.get("similarity_score", 0) >= chunk.get("similarity_score", 0) - 0.03
        ):
            chunk = row_chunk

        relevant_chunks.append(chunk)

        if len(relevant_chunks) >= EVIDENCE_TOP_CHUNKS:
            break

    return relevant_chunks


def build_evidence_location(relevant_chunks):
    if not relevant_chunks:
        return ""

    pdf_pages = sorted({
        chunk.get("page_number")
        for chunk in relevant_chunks
        if chunk.get("page_number") is not None
    })
    if pdf_pages:
        if len(pdf_pages) == 1:
            return f"Page: {pdf_pages[0]}"
        return f"Pages: {pdf_pages[0]}-{pdf_pages[-1]}"

    sheet_names = {
        chunk.get("sheet_name")
        for chunk in relevant_chunks
        if chunk.get("sheet_name")
    }
    row_numbers = sorted({
        chunk.get("row_number")
        for chunk in relevant_chunks
        if chunk.get("row_number") is not None
    })
    columns = [
        chunk.get("column_name")
        for chunk in relevant_chunks
        if chunk.get("column_name")
    ]

    if sheet_names and row_numbers:
        sheet_name = sorted(sheet_names)[0]
        rows_text = format_compact_values(row_numbers)
        columns_text = ", ".join(sorted(set(columns)))
        if columns_text:
            return f"Sheet: {sheet_name}, Rows: {rows_text}, Columns: {columns_text}"
        return f"Sheet: {sheet_name}, Rows: {rows_text}"

    return relevant_chunks[0].get("location", "")


def format_compact_values(values, max_items=5):
    unique_values = list(dict.fromkeys(values))
    if len(unique_values) <= max_items:
        return ", ".join(str(value) for value in unique_values)

    preview = ", ".join(str(value) for value in unique_values[:max_items])
    return f"{preview} (+{len(unique_values) - max_items} more)"


def build_evidence_snippets(useful_chunks, max_total_chars=EVIDENCE_PROMPT_CHARS):
    snippets = []
    total_length = 0

    for index, chunk in enumerate(useful_chunks[:EVIDENCE_TOP_CHUNKS], start=1):
        location = clean_text(chunk.get("location")) or build_evidence_location([chunk])
        text_value = truncate_text(chunk.get("text"), EVIDENCE_SNIPPET_CHARS)
        if not text_value:
            continue

        snippet = f"[{index}] {location}: {text_value}"
        if total_length + len(snippet) > max_total_chars:
            remaining_chars = max_total_chars - total_length
            if remaining_chars <= 80:
                break
            snippet = truncate_text(snippet, remaining_chars)

        snippets.append(snippet)
        total_length += len(snippet)

    return "\n".join(snippets)


def clamp_confidence(value):
    try:
        confidence = int(value)
    except (TypeError, ValueError):
        return 50

    return max(0, min(100, confidence))


def parse_model_json_response(content):
    cleaned_content = clean_text(content)

    try:
        return json.loads(cleaned_content)
    except json.JSONDecodeError:
        json_match = re.search(r"\{.*\}", cleaned_content, re.DOTALL)
        if not json_match:
            raise HTTPException(
                status_code=502,
                detail="Qwen returned a response that could not be parsed as JSON."
            )

        try:
            return json.loads(json_match.group(0))
        except json.JSONDecodeError as exc:
            raise HTTPException(
                status_code=502,
                detail="Qwen returned malformed JSON."
            ) from exc


def call_ollama_chat(prompt):
    payload = {
        "model": AI_MODEL_NAME,
        "stream": False,
        "messages": [
            {
                "role": "system",
                "content": (
                    "You are an NDMO compliance evidence assessor. "
                    "Return only valid JSON. Do not include markdown."
                )
            },
            {
                "role": "user",
                "content": prompt
            }
        ],
        "options": {
            "temperature": 0.1,
            "num_thread": OLLAMA_NUM_THREAD,
            "num_ctx": OLLAMA_NUM_CTX,
            "num_predict": OLLAMA_NUM_PREDICT
        }
    }

    data = json.dumps(payload).encode("utf-8")
    request = urlrequest.Request(
        f"{OLLAMA_BASE_URL}/api/chat",
        data=data,
        headers={"Content-Type": "application/json"},
        method="POST"
    )

    try:
        with urlrequest.urlopen(request, timeout=OLLAMA_TIMEOUT_SECONDS) as response:
            response_payload = json.loads(response.read().decode("utf-8"))
    except HTTPError as exc:
        raise HTTPException(
            status_code=502,
            detail=f"Ollama request failed with status {exc.code}."
        ) from exc
    except URLError as exc:
        raise HTTPException(
            status_code=503,
            detail="Could not connect to Ollama. Start Ollama and pull the selected Qwen model first."
        ) from exc

    return response_payload.get("message", {}).get("content", "")


def normalize_response_language(response_language):
    language = clean_text(response_language).lower()
    if language.startswith("ar"):
        return "Arabic"
    return "English"


def normalize_answer_label(answer_value):
    normalized = clean_text(answer_value).lower()
    answer_aliases = {
        "yes": "yes",
        "y": "yes",
        "نعم": "yes",
        "partial": "partial",
        "partially": "partial",
        "جزئي": "partial",
        "جزئية": "partial",
        "no": "no",
        "n": "no",
        "لا": "no",
    }
    return answer_aliases.get(normalized, normalized)


def infer_answer_with_qwen(question_text, useful_chunks, response_language="English", domain_context=None):
    response_language = normalize_response_language(response_language)
    domain_context = domain_context or {"names": []}
    domain_name = domain_context.get("names", [""])[0] if domain_context.get("names") else ""
    evidence_text = build_evidence_snippets(useful_chunks)
    evidence_location = build_evidence_location(useful_chunks[:EVIDENCE_TOP_CHUNKS])
    reason_instruction = (
        "Write the reason in clear Modern Standard Arabic."
        if response_language == "Arabic"
        else "Write the reason in clear English."
    )

    prompt = f"""
Assess whether the evidence supports this NDMO requirement.

Requirement:
{question_text}

Current NDMO domain:
{domain_name or "Not specified"}

Evidence:
{evidence_text}

Return JSON only with this schema:
{{
  "ai_answer": "yes" | "partial" | "no",
  "confidence": 0-100,
  "reason": "short reason in {response_language}"
}}

Rules:
- yes means the evidence clearly proves the requirement.
- partial means the evidence is related but incomplete, draft, missing approval, or indirect.
- no means the evidence does not support the requirement.
- The evidence snippets were selected by multilingual semantic embeddings, not keyword matching.
- Use only the numbered evidence snippets above and never invent missing facts.
- Column names or headers alone are not proof; require populated values or explicit document content.
- Empty, N/A, draft, pending, and unapproved fields cannot support a yes decision.
- For an owner requirement, require an actual assigned owner or accountable entity.
- For a measurement requirement, require an actual metric or review value and a stated cadence for yes.
- For a policy requirement, require an actual policy/procedure and approval evidence for yes.
- Evidence must belong to the current NDMO domain, or clearly apply to all domains.
- Do not infer support from rows for a different domain.
- Keep JSON keys in English.
- {reason_instruction}
""".strip()

    content = call_ollama_chat(prompt)
    model_result = parse_model_json_response(content)
    answer = clean_text(model_result.get("ai_answer") or model_result.get("answer")).lower()

    if answer not in ["yes", "partial", "no"]:
        raise HTTPException(
            status_code=502,
            detail="Qwen returned an invalid assessment label. No fallback decision was created."
        )

    return {
        "ai_answer": answer,
        "confidence": clamp_confidence(model_result.get("confidence")),
        "reason": clean_text(model_result.get("reason")) or (
            "قيّم Qwen الدليل المرفوع مقابل متطلب NDMO."
            if response_language == "Arabic"
            else "Qwen assessed the uploaded evidence against the NDMO requirement."
        ),
        "evidence_text": evidence_text[:900],
        "evidence_location": evidence_location
    }


def infer_answer_from_evidence(question_text, relevant_chunks, response_language="English", domain_context=None):
    response_language = normalize_response_language(response_language)
    domain_context = domain_context or {"names": []}
    useful_chunks = [
        chunk for chunk in relevant_chunks
        if chunk.get("retrieval_method") == "semantic"
        or chunk.get("similarity_score", 0) > 0
    ]

    if not useful_chunks:
        return {
            "ai_answer": "no",
            "confidence": 25,
            "reason": (
                "لم يتم العثور على نص دليل مرتبط بهذا السؤال."
                if response_language == "Arabic"
                else "No relevant evidence text was found for this question."
            ),
            "evidence_text": "",
            "evidence_location": ""
        }

    if AI_PROVIDER == "ollama":
        return infer_answer_with_qwen(
            question_text,
            useful_chunks,
            response_language,
            domain_context=domain_context
        )

    proof_chunks = [
        chunk for chunk in useful_chunks
        if chunk.get("chunk_level") != "header"
    ]
    if not proof_chunks:
        return {
            "ai_answer": "no",
            "confidence": 70,
            "reason": (
                "عناوين الأعمدة وحدها لا تثبت المتطلب، ولم تظهر قيم فعلية تدعمه."
                if response_language == "Arabic"
                else "Column headers alone do not prove the requirement, and no supporting values were found."
            ),
            "evidence_text": "",
            "evidence_location": ""
        }

    best_score = max(chunk.get("similarity_score", 0) for chunk in proof_chunks)
    evidence_text = build_evidence_snippets(proof_chunks, max_total_chars=900)
    evidence_location = build_evidence_location(proof_chunks[:EVIDENCE_TOP_CHUNKS])
    proof_search_text = " ; ".join(get_chunk_search_text(chunk) for chunk in proof_chunks)
    question_intent = detect_question_intent(question_text)

    if question_intent == "policy":
        document_values = get_values_from_chunks(proof_chunks, POLICY_DOCUMENT_LABELS)
        approval_values = get_values_from_chunks(proof_chunks, APPROVAL_STATUS_LABELS)
        has_document = any(
            is_meaningful_evidence_value(value)
            and text_contains_any_term(value, ["policy", "procedure", "standard", "سياسة", "اجراء", "معيار"])
            for value in document_values
        )
        has_policy_phrase = text_contains_any_term(
            proof_search_text,
            ["policy", "procedure", "documented", "سياسة", "اجراء", "موثق"]
        )
        has_policy_container = any(
            chunk.get("file_type") in {".pdf", ".txt"}
            or text_contains_any_term(
                " ".join([
                    clean_text(chunk.get("sheet_name")),
                    clean_text(chunk.get("column_name")),
                ]),
                ["policy", "procedure", "document", "standard", "سياسة", "اجراء", "وثيقة", "معيار"]
            )
            for chunk in proof_chunks
        )
        has_policy_text = best_score >= 0.24 and has_policy_phrase and has_policy_container
        is_approved = any_normalized_value_contains(
            approval_values,
            {"approved", "accepted", "valid", "active", "معتمد", "معتمدة", "مقبول", "مقبولة", "ساري", "سارية"}
        ) or text_contains_any_term(
            proof_search_text,
            ["approved by", "approval date", "معتمد من", "تاريخ الاعتماد"]
        )
        is_incomplete = any_normalized_value_contains(
            approval_values,
            {"draft", "pending", "expired", "rejected", "مسودة", "قيد المراجعة", "منتهي", "مرفوض"}
        )
        has_implementation = has_domain_implementation_signal(proof_chunks, domain_context)

        if (has_document or has_policy_text) and is_approved and not is_incomplete:
            ai_answer = "yes"
            confidence = min(95, int(82 + (best_score * 20)))
            reason = (
                "يتضمن الملف سياسة أو إجراءً موثقًا مع ما يثبت اعتماده لهذا المجال."
                if response_language == "Arabic"
                else "The file contains a documented policy or procedure with approval evidence for this domain."
            )
        elif has_document or has_policy_text or has_implementation:
            ai_answer = "partial"
            confidence = min(88, int(68 + (best_score * 24)))
            reason = (
                "يوضح الملف تطبيقًا أو محتوى مرتبطًا بالمجال، لكنه لا يتضمن سياسة أو إجراءً موثقًا مع حالة اعتماد مكتملة."
                if response_language == "Arabic"
                else "The file shows related implementation or content, but it does not include a documented policy or procedure with complete approval evidence."
            )
        else:
            ai_answer = "no"
            confidence = min(86, int(70 + (best_score * 20)))
            reason = (
                "لم يظهر في الملف دليل على سياسة أو إجراء موثق ومعتمد يخص هذا المجال."
                if response_language == "Arabic"
                else "No evidence of a documented and approved policy or procedure for this domain was found in the file."
            )

    elif question_intent == "owner":
        owner_values = get_values_from_chunks(proof_chunks, OWNER_LABELS)
        assigned_owners = [
            value for value in owner_values
            if is_meaningful_evidence_value(value)
        ]

        if assigned_owners:
            ai_answer = "yes"
            confidence = min(95, int(84 + (best_score * 18)))
            reason = (
                "يتضمن الملف اسم مالك أو مسؤول معيّن فعليًا لهذا المجال."
                if response_language == "Arabic"
                else "The file contains an actual assigned owner or accountable person for this domain."
            )
        else:
            ai_answer = "no"
            confidence = min(88, int(72 + (best_score * 18)))
            reason = (
                "لم يظهر في القيم المستخرجة من الملف اسم مالك أو مسؤول معيّن لهذا المجال."
                if response_language == "Arabic"
                else "No assigned owner or accountable person for this domain was found in the extracted file values."
            )

    elif question_intent == "measurement":
        measurement_values = get_values_from_chunks(proof_chunks, MEASUREMENT_LABELS)
        has_measurement_value = any(
            is_meaningful_evidence_value(value, allow_boolean=True)
            for value in measurement_values
        )
        has_measurement_text = text_contains_any_term(
            proof_search_text,
            ["quality score", "metric", "measurement", "kpi", "review", "improvement", "درجة الجودة", "مؤشر", "قياس", "مراجعة", "تحسين"]
        )
        has_cadence = text_contains_any_term(
            proof_search_text,
            ["periodic", "monthly", "quarterly", "semiannual", "annual", "weekly", "دوري", "شهري", "ربع سنوي", "نصف سنوي", "سنوي", "اسبوعي"]
        )

        if (has_measurement_value or has_measurement_text) and has_cadence:
            ai_answer = "yes"
            confidence = min(94, int(82 + (best_score * 18)))
            reason = (
                "يتضمن الملف قياسًا أو مراجعة موثقة مع دورية واضحة لهذا المجال."
                if response_language == "Arabic"
                else "The file contains documented measurement or review evidence with a clear cadence for this domain."
            )
        elif has_measurement_value:
            ai_answer = "partial"
            confidence = min(84, int(66 + (best_score * 22)))
            reason = (
                "يتضمن الملف مؤشرًا أو مراجعة مرتبطة بالمجال، لكنه لا يثبت قياسًا وتحسينًا دوريًا مكتملًا."
                if response_language == "Arabic"
                else "The file includes a related metric or review, but it does not prove a complete periodic measurement and improvement process."
            )
        else:
            ai_answer = "no"
            confidence = min(88, int(72 + (best_score * 18)))
            reason = (
                "لم يظهر في الملف دليل على قياس هذا المجال وتحسينه وفق دورية محددة."
                if response_language == "Arabic"
                else "No evidence of measuring and improving this domain on a defined cadence was found in the file."
            )

    else:
        question_tokens = set(tokenize(question_text))
        evidence_tokens = set(tokenize(proof_search_text))
        direct_matches = len(question_tokens & evidence_tokens)
        formal_terms = {
            "approved", "policy", "procedure", "standard", "owner", "documented",
            "implemented", "process", "governance", "review", "classification",
            "quality", "controls", "approval", "اجراء", "سياسة", "معتمد"
        }
        formal_matches = len(evidence_tokens & formal_terms)

        if best_score >= 0.45 and direct_matches >= 3 and formal_matches >= 3:
            ai_answer = "yes"
            confidence = min(93, int(76 + (best_score * 30)))
            reason = (
                "يتضمن الملف دليلًا مباشرًا ورسميًا يدعم المتطلب."
                if response_language == "Arabic"
                else "The file contains direct, formal evidence supporting the requirement."
            )
        elif best_score >= 0.24 and direct_matches >= 2:
            ai_answer = "partial"
            confidence = min(82, int(62 + (best_score * 28)))
            reason = (
                "يتضمن الملف دليلًا مرتبطًا، لكنه غير كافٍ لإثبات المتطلب بالكامل."
                if response_language == "Arabic"
                else "The file contains related evidence, but it is insufficient to fully prove the requirement."
            )
        else:
            ai_answer = "no"
            confidence = min(86, int(68 + (best_score * 20)))
            reason = (
                "الدليل المستخرج ضعيف الصلة ولا يثبت المتطلب بوضوح."
                if response_language == "Arabic"
                else "The extracted evidence is weakly related and does not clearly prove the requirement."
            )

    return {
        "ai_answer": ai_answer,
        "confidence": confidence,
        "reason": reason,
        "evidence_text": evidence_text,
        "evidence_location": evidence_location
    }


def get_file_chunks(file_path):
    suffix = Path(file_path).suffix.lower()

    if suffix in [".xlsx", ".xlsm", ".xltx", ".xltm"]:
        extracted_items = extract_text_from_excel(file_path)
    elif suffix == ".pdf":
        extracted_items = extract_text_from_pdf(file_path)
    elif suffix == ".csv":
        extracted_items = extract_text_from_csv(file_path)
    elif suffix == ".txt":
        extracted_items = extract_text_from_plain_file(file_path)
    else:
        raise HTTPException(
            status_code=400,
            detail="Unsupported evidence file type. Please upload Excel, PDF, CSV, or TXT evidence."
        )

    return chunk_extracted_text(extracted_items)


def build_semantic_embedding_text(chunk):
    return truncate_text(get_chunk_search_text(chunk), 6000)


def call_ollama_embeddings(text_values):
    inputs = [normalize_evidence_text(value) for value in text_values if normalize_evidence_text(value)]
    if not inputs:
        return []

    payload = {
        "model": EMBEDDING_MODEL_NAME,
        "input": inputs,
        "truncate": True,
    }
    request = urlrequest.Request(
        f"{OLLAMA_BASE_URL}/api/embed",
        data=json.dumps(payload).encode("utf-8"),
        headers={"Content-Type": "application/json"},
        method="POST",
    )

    try:
        with urlrequest.urlopen(request, timeout=EMBEDDING_TIMEOUT_SECONDS) as response:
            response_payload = json.loads(response.read().decode("utf-8"))
    except HTTPError as exc:
        raise HTTPException(
            status_code=502,
            detail=f"The local embedding model request failed with status {exc.code}."
        ) from exc
    except URLError as exc:
        raise HTTPException(
            status_code=503,
            detail="Could not connect to the local semantic embedding model."
        ) from exc

    embeddings = response_payload.get("embeddings") or []
    if len(embeddings) != len(inputs) or any(not embedding for embedding in embeddings):
        raise HTTPException(
            status_code=502,
            detail="The local embedding model returned an incomplete semantic index."
        )

    return embeddings


def semantic_vector_similarity(vector_a, vector_b):
    if not vector_a or not vector_b or len(vector_a) != len(vector_b):
        return 0
    return sum(float(left) * float(right) for left, right in zip(vector_a, vector_b))


def select_semantic_chunks(extracted_chunks):
    return [
        chunk for chunk in extracted_chunks
        if chunk.get("chunk_level") == "row" or not chunk.get("chunk_level")
    ]


def prepare_semantic_chunks(extracted_chunks):
    semantic_chunks = [chunk.copy() for chunk in select_semantic_chunks(extracted_chunks)]
    if not semantic_chunks:
        return []

    batch_size = max(1, EMBEDDING_BATCH_SIZE)
    for start in range(0, len(semantic_chunks), batch_size):
        batch = semantic_chunks[start:start + batch_size]
        batch_texts = [build_semantic_embedding_text(chunk) for chunk in batch]
        batch_embeddings = call_ollama_embeddings(batch_texts)
        for chunk, embedding in zip(batch, batch_embeddings):
            chunk["_semantic_embedding"] = embedding

    return semantic_chunks


def get_analysis_chunks(file_path):
    if not SEMANTIC_RETRIEVAL_ENABLED:
        return get_file_chunks(file_path)

    file_stat = os.stat(file_path)
    file_digest = hashlib.sha256()
    with open(file_path, "rb") as source:
        for block in iter(lambda: source.read(1024 * 1024), b""):
            file_digest.update(block)
    cache_key = (
        Path(file_path).suffix.lower(),
        file_stat.st_size,
        file_digest.hexdigest(),
        EMBEDDING_MODEL_NAME,
    )
    cached_chunks = SEMANTIC_CHUNK_CACHE.get(cache_key)
    if cached_chunks is not None:
        return cached_chunks

    prepared_chunks = prepare_semantic_chunks(get_file_chunks(file_path))
    if not prepared_chunks:
        raise HTTPException(
            status_code=400,
            detail="No evidence rows or document text were available for semantic analysis."
        )

    SEMANTIC_CHUNK_CACHE[cache_key] = prepared_chunks
    while len(SEMANTIC_CHUNK_CACHE) > max(1, SEMANTIC_CACHE_FILES):
        oldest_key = next(iter(SEMANTIC_CHUNK_CACHE))
        SEMANTIC_CHUNK_CACHE.pop(oldest_key, None)

    return prepared_chunks


def build_semantic_question_text(question_text, domain_context):
    domain_names = " / ".join(domain_context.get("names", []))
    return "\n".join([
        f"NDMO domain / مجال NDMO: {domain_names}",
        f"Compliance requirement / متطلب الامتثال: {question_text}",
        "Retrieve evidence that directly proves or disproves this requirement.",
    ])


def resolve_evidence_file_path(evidence):
    file_path = evidence.get("file_path")
    if not file_path:
        raise HTTPException(status_code=400, detail="Evidence file path is missing")

    if os.path.exists(file_path):
        return file_path

    possible_path = os.path.join(EVIDENCE_UPLOAD_DIR, evidence.get("file_name", ""))
    if os.path.exists(possible_path):
        return possible_path

    raise HTTPException(status_code=404, detail="Evidence file was not found on disk")


def score_chunks_semantically(question_text, extracted_chunks, domain_context=None):
    domain_context = domain_context or {"names": []}
    query_text = build_semantic_question_text(question_text, domain_context)
    query_embeddings = call_ollama_embeddings([query_text])
    if not query_embeddings:
        raise HTTPException(status_code=502, detail="The semantic query embedding was not generated.")

    query_embedding = query_embeddings[0]
    candidate_limit = max(80, EVIDENCE_CANDIDATE_CHUNKS * 25)
    candidate_heap = []

    for index, chunk in enumerate(extracted_chunks):
        chunk_embedding = chunk.get("_semantic_embedding")
        if not chunk_embedding:
            continue

        score = max(0, round(semantic_vector_similarity(query_embedding, chunk_embedding), 6))
        scored_chunk = chunk.copy()
        scored_chunk["similarity_score"] = score
        scored_chunk["retrieval_method"] = "semantic"
        rank = (score, index)
        entry = (*rank, scored_chunk)
        if len(candidate_heap) < candidate_limit:
            heapq.heappush(candidate_heap, entry)
        elif rank > candidate_heap[0][:2]:
            heapq.heapreplace(candidate_heap, entry)

    if not candidate_heap:
        raise HTTPException(
            status_code=502,
            detail="The semantic evidence index is empty. Text matching was not used as a fallback."
        )

    return [
        entry[2]
        for entry in sorted(candidate_heap, key=lambda item: item[:2], reverse=True)
    ]


def score_chunks_for_question(question_text, extracted_chunks, domain_context=None):
    domain_context = domain_context or {"phrases": set(), "tokens": set()}
    if SEMANTIC_RETRIEVAL_ENABLED:
        return score_chunks_semantically(
            question_text,
            extracted_chunks,
            domain_context=domain_context,
        )

    has_domain_matches = any(
        get_chunk_domain_alignment(chunk, domain_context) == "match"
        for chunk in extracted_chunks
    )
    candidate_limit = max(80, EVIDENCE_CANDIDATE_CHUNKS * 25)
    candidate_heap = []

    for index, chunk in enumerate(extracted_chunks):
        if has_domain_matches and get_chunk_domain_alignment(chunk, domain_context) != "match":
            continue
        score = score_chunk_for_question(
            question_text,
            chunk,
            domain_context=domain_context
        )
        scored_chunk = chunk.copy()
        scored_chunk["similarity_score"] = score
        rank = (score, 1 if chunk.get("chunk_level") in ["row", "header"] else 0, index)
        entry = (*rank, scored_chunk)
        if len(candidate_heap) < candidate_limit:
            heapq.heappush(candidate_heap, entry)
        elif rank > candidate_heap[0][:3]:
            heapq.heapreplace(candidate_heap, entry)

    return [
        entry[3]
        for entry in sorted(candidate_heap, key=lambda item: item[:3], reverse=True)
    ]


def answer_to_score(answer_value):
    score_map = {
        "yes": 100,
        "partial": 50,
        "no": 0
    }
    return score_map.get(clean_text(answer_value).lower(), 0)


def build_ai_score_summary(results):
    total_questions = len(results)
    yes_count = sum(1 for result in results if result["ai_answer"] == "yes")
    partial_count = sum(1 for result in results if result["ai_answer"] == "partial")
    no_count = sum(1 for result in results if result["ai_answer"] == "no")
    total_score = sum(answer_to_score(result["ai_answer"]) for result in results)

    if total_questions:
        score = round(total_score / total_questions, 2)
    else:
        score = 0

    return {
        "total_questions": total_questions,
        "yes_count": yes_count,
        "partial_count": partial_count,
        "no_count": no_count,
        "domain_score": score,
        "ndi_score": score
    }


def save_document_chunks(cur, question_id, evidence_id, chunks):
    for chunk in chunks:
        cur.execute("""
        INSERT INTO ndmo_document_chunks
            (
                evidence_id,
                question_id,
                chunk_text,
                file_name,
                file_type,
                page_number,
                sheet_name,
                row_number,
                column_name,
                similarity_score
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
        """, (
            evidence_id,
            question_id,
            chunk.get("text", ""),
            chunk.get("file_name"),
            chunk.get("file_type"),
            chunk.get("page_number"),
            chunk.get("sheet_name"),
            chunk.get("row_number"),
            chunk.get("column_name"),
            chunk.get("similarity_score")
        ))


def save_ai_assessment_result(cur, question_id, evidence_id, ai_result):
    cur.execute("""
        INSERT INTO ndmo_ai_assessment_results
        (
            question_id,
            evidence_id,
            ai_answer,
            confidence,
            reason,
            evidence_text,
            evidence_location
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s)
        RETURNING id, created_at;
    """, (
        question_id,
        evidence_id,
        ai_result["ai_answer"],
        ai_result["confidence"],
        ai_result["reason"],
        ai_result["evidence_text"],
        ai_result["evidence_location"]
    ))

    return cur.fetchone()


def analyze_question_with_chunks(
    question,
    evidence_id,
    extracted_chunks,
    cur,
    save_all_chunks=False,
    response_language="English"
):
    question_text = get_question_text(question)
    if not question_text:
        return None

    domain_context = build_question_domain_context(question)
    scored_chunks = score_chunks_for_question(
        question_text,
        extracted_chunks,
        domain_context=domain_context
    )
    relevant_chunks = find_relevant_evidence(
        question_text,
        scored_chunks,
        domain_context=domain_context
    )
    ai_result = infer_answer_from_evidence(
        question_text,
        relevant_chunks,
        response_language=response_language,
        domain_context=domain_context
    )

    chunks_to_save = scored_chunks if save_all_chunks else relevant_chunks
    save_document_chunks(cur, question["id"], evidence_id, chunks_to_save)
    saved_result = save_ai_assessment_result(cur, question["id"], evidence_id, ai_result)

    return {
        "result_id": saved_result["id"],
        "question_id": question["id"],
        "domain_id": question.get("domain_id"),
        "question_code": question.get("question_code"),
        "question_text": question_text,
        "ai_answer": ai_result["ai_answer"],
        "score": answer_to_score(ai_result["ai_answer"]),
        "confidence": ai_result["confidence"],
        "reason": ai_result["reason"],
        "evidence_text": ai_result["evidence_text"],
        "evidence_location": ai_result["evidence_location"],
        "created_at": str(saved_result["created_at"])
    }


def save_domain_assessment_run(cur, domain_id, evidence_id, summary):
    cur.execute("""
        INSERT INTO ndmo_ai_domain_assessment_runs
        (
            domain_id,
            evidence_id,
            total_questions,
            yes_count,
            partial_count,
            no_count,
            domain_score,
            ndi_score,
            model_name
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        RETURNING id, created_at;
    """, (
        domain_id,
        evidence_id,
        summary["total_questions"],
        summary["yes_count"],
        summary["partial_count"],
        summary["no_count"],
        summary["domain_score"],
        summary["ndi_score"],
        AI_MODEL_NAME
    ))

    return cur.fetchone()


def limit_questions_for_analysis(questions, max_questions=None, question_offset=0):
    total_questions = len(questions)

    try:
        offset = int(question_offset)
    except (TypeError, ValueError):
        offset = 0

    offset = max(0, min(offset, total_questions))

    try:
        question_limit = int(max_questions)
    except (TypeError, ValueError):
        question_limit = MAX_AI_QUESTIONS if AI_PROVIDER == "ollama" else total_questions

    if question_limit <= 0:
        question_limit = MAX_AI_QUESTIONS if AI_PROVIDER == "ollama" else total_questions

    limited_questions = questions[offset:offset + question_limit]
    next_offset = offset + len(limited_questions)

    return limited_questions, {
        "limited": next_offset < total_questions,
        "analyzed_questions": len(limited_questions),
        "question_offset": offset,
        "next_offset": next_offset,
        "batch_size": question_limit,
        "total_available_questions": total_questions
    }


def analyze_questions_against_evidence(
    questions,
    evidence,
    save_all_chunks=False,
    response_language="English",
    max_questions=None,
    question_offset=0
):
    questions, analysis_limit = limit_questions_for_analysis(
        questions,
        max_questions,
        question_offset=question_offset
    )
    file_path = resolve_evidence_file_path(evidence)
    extracted_chunks = get_analysis_chunks(file_path)
    if not extracted_chunks:
        raise HTTPException(status_code=400, detail="No readable text was found in the evidence file")

    results = []
    conn = get_connection()
    cur = conn.cursor()

    for question in questions:
        result = analyze_question_with_chunks(
            question,
            evidence["id"],
            extracted_chunks,
            cur,
            save_all_chunks=save_all_chunks,
            response_language=response_language
        )
        if result is not None:
            results.append(result)

    conn.commit()
    cur.close()
    conn.close()

    return results, analysis_limit


async def save_uploaded_evidence_file(question_id, evidence_type, file, user=None):
    original_name, file_path = await persist_evidence_upload(file)

    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        INSERT INTO ndmo_evidence
        (question_id, file_name, file_path, evidence_type, status)
        VALUES (%s, %s, %s, %s, %s)
        RETURNING id, question_id, file_name, file_path, evidence_type, status, uploaded_at;
    """, (question_id, original_name, file_path, evidence_type, "Uploaded"))

    evidence = cur.fetchone()
    conn.commit()
    cur.close()
    conn.close()

    if user:
        write_audit(user, "upload_evidence", "ndmo_evidence", evidence["id"], {"file_name": original_name, "evidence_type": evidence_type})

    return evidence


@app.get("/api/ndmo/domains")
def get_ndmo_domains():
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT id, name_ar, name_en, description, order_number
        FROM ndmo_domains
        ORDER BY order_number;
    """)

    domains = cur.fetchall()
    cur.close()
    conn.close()

    return {"domains": domains}


@app.get("/api/ndmo/evidence-files")
def get_ndmo_evidence_files():
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT DISTINCT ON (file_name)
            id, file_name, evidence_type, status, uploaded_at
        FROM ndmo_evidence
        ORDER BY file_name, uploaded_at DESC, id DESC;
    """)
    files = cur.fetchall()
    cur.close()
    conn.close()
    return {"files": files}


@app.get("/api/ndmo/questions")
def get_ndmo_questions():
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT 
            q.id,
            q.question_code,
            q.question_text_ar,
            q.question_text_en,
            q.evidence_required,
            q.max_score,
            d.id AS domain_id,
            d.name_ar AS domain_name_ar,
            d.name_en AS domain_name_en,
            d.order_number
        FROM ndmo_questions q
        JOIN ndmo_domains d ON q.domain_id = d.id
        ORDER BY d.order_number, q.id;
    """)

    questions = cur.fetchall()
    cur.close()
    conn.close()

    return {"questions": questions}


@app.post("/api/ndmo/answers")
def submit_ndmo_answer(
    request: Request,
    question_id: int = Form(...),
    answer_value: str = Form(...),
    notes: str = Form("")
):
    score_map = {
        "yes": 100,
        "partial": 50,
        "no": 0,
        "not_applicable": 0
    }

    normalized_answer = answer_value.lower()
    if normalized_answer not in score_map:
        raise HTTPException(status_code=422, detail="Invalid answer value")
    score = score_map[normalized_answer]

    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        INSERT INTO ndmo_answers (question_id, answer_value, score, notes)
        VALUES (%s, %s, %s, %s)
        RETURNING id, question_id, answer_value, score, notes, created_at;
    """, (question_id, answer_value, score, notes))

    answer = cur.fetchone()
    write_audit(
        request.state.current_user,
        "save_manual_answer",
        "ndmo_question",
        question_id,
        {"answer": normalized_answer, "score": score},
        conn=conn,
    )
    conn.commit()

    cur.close()
    conn.close()

    return {
        "message": "Answer saved successfully",
        "answer": answer
    }


@app.post("/api/ndmo/evidence")
async def upload_ndmo_evidence(
    request: Request,
    question_id: int = Form(...),
    evidence_type: str = Form("Document"),
    file: UploadFile = File(...)
):
    evidence = await save_uploaded_evidence_file(question_id, evidence_type, file, request.state.current_user)

    return {
        "message": "Evidence uploaded successfully",
        "evidence": evidence
    }


@app.post("/api/ndmo/upload-and-analyze-domain")
async def upload_and_analyze_ndmo_domain_evidence(
    request: Request,
    domain_id: int = Form(...),
    evidence_type: str = Form("Document"),
    response_language: str = Form("English"),
    max_questions: int = Form(MAX_AI_QUESTIONS),
    question_offset: int = Form(0),
    file: UploadFile = File(...)
):
    create_ai_assessment_tables()

    domain = get_domain_by_id(domain_id)
    if domain is None:
        raise HTTPException(status_code=404, detail="Domain not found")

    questions = get_questions_by_domain_id(domain_id)
    if not questions:
        raise HTTPException(status_code=404, detail="No questions found for this domain")

    evidence = await save_uploaded_evidence_file(
        questions[0]["id"],
        evidence_type,
        file,
        request.state.current_user,
    )

    results, analysis_limit = await run_in_threadpool(
        analyze_questions_against_evidence,
        questions,
        evidence,
        save_all_chunks=False,
        response_language=response_language,
        max_questions=max_questions,
        question_offset=question_offset
    )
    summary = build_ai_score_summary(results)

    conn = get_connection()
    cur = conn.cursor()
    saved_run = save_domain_assessment_run(cur, domain_id, evidence["id"], summary)
    conn.commit()
    cur.close()
    conn.close()

    return {
        "message": "Evidence uploaded and domain analyzed successfully",
        "run_id": saved_run["id"],
        "model_name": AI_MODEL_NAME,
        "base_model_name": AI_BASE_MODEL_NAME,
        "domain": domain,
        "evidence": evidence,
        "analysis_limit": analysis_limit,
        "summary": summary,
        "questions": results,
        "created_at": str(saved_run["created_at"])
    }


@app.post("/api/ndmo/upload-and-analyze-all")
async def upload_and_analyze_all_ndmo_evidence(
    request: Request,
    evidence_type: str = Form("Document"),
    response_language: str = Form("English"),
    max_questions: int = Form(MAX_AI_QUESTIONS),
    question_offset: int = Form(0),
    file: UploadFile = File(...)
):
    create_ai_assessment_tables()

    questions = get_all_ndmo_questions()
    if not questions:
        raise HTTPException(status_code=404, detail="No NDMO questions found")

    evidence = await save_uploaded_evidence_file(
        questions[0]["id"],
        evidence_type,
        file,
        request.state.current_user,
    )

    results, analysis_limit = await run_in_threadpool(
        analyze_questions_against_evidence,
        questions,
        evidence,
        save_all_chunks=False,
        response_language=response_language,
        max_questions=max_questions,
        question_offset=question_offset
    )

    domain_summaries = []
    conn = get_connection()
    cur = conn.cursor()

    domain_ids = sorted({
        result["domain_id"]
        for result in results
        if result.get("domain_id") is not None
    })

    for domain_id in domain_ids:
        domain = get_domain_by_id(domain_id)
        domain_results = [
            result for result in results
            if result.get("domain_id") == domain_id
        ]
        summary = build_ai_score_summary(domain_results)
        saved_run = save_domain_assessment_run(cur, domain_id, evidence["id"], summary)

        domain_summaries.append({
            "run_id": saved_run["id"],
            "domain": domain,
            "summary": summary,
            "questions": domain_results,
            "created_at": str(saved_run["created_at"])
        })

    overall_summary = build_ai_score_summary(results)
    conn.commit()
    cur.close()
    conn.close()

    return {
        "message": "Evidence uploaded and all NDMO questions analyzed successfully",
        "model_name": AI_MODEL_NAME,
        "base_model_name": AI_BASE_MODEL_NAME,
        "evidence": evidence,
        "analysis_limit": analysis_limit,
        "overall_ndi_score": overall_summary["ndi_score"],
        "overall_summary": overall_summary,
        "domains": domain_summaries
    }


@app.post("/api/ndmo/analyze-evidence")
def analyze_ndmo_evidence(
    question_id: int = Form(...),
    evidence_id: int = Form(...),
    response_language: str = Form("English")
):
    create_ai_assessment_tables()

    question = get_question_by_id(question_id)
    if question is None:
        raise HTTPException(status_code=404, detail="Question not found")

    evidence = get_evidence_by_id(evidence_id)
    if evidence is None:
        raise HTTPException(status_code=404, detail="Evidence not found")

    results, analysis_limit = analyze_questions_against_evidence(
        [question],
        evidence,
        save_all_chunks=True,
        response_language=response_language
    )

    if not results:
        raise HTTPException(status_code=400, detail="Question text is empty")

    result = results[0]

    return {
        "message": "Evidence analyzed successfully",
        "result_id": result["result_id"],
        "question_id": question_id,
        "evidence_id": evidence_id,
        "ai_answer": result["ai_answer"],
        "confidence": result["confidence"],
        "reason": result["reason"],
        "evidence_text": result["evidence_text"],
        "evidence_location": result["evidence_location"],
        "analysis_limit": analysis_limit,
        "created_at": result["created_at"]
    }


@app.post("/api/ndmo/analyze-domain-evidence")
def analyze_ndmo_domain_evidence(
    domain_id: int = Form(...),
    evidence_id: int = Form(...),
    response_language: str = Form("English"),
    max_questions: int = Form(MAX_AI_QUESTIONS),
    question_offset: int = Form(0)
):
    create_ai_assessment_tables()

    domain = get_domain_by_id(domain_id)
    if domain is None:
        raise HTTPException(status_code=404, detail="Domain not found")

    evidence = get_evidence_by_id(evidence_id)
    if evidence is None:
        raise HTTPException(status_code=404, detail="Evidence not found")

    questions = get_questions_by_domain_id(domain_id)
    if not questions:
        raise HTTPException(status_code=404, detail="No questions found for this domain")

    results, analysis_limit = analyze_questions_against_evidence(
        questions,
        evidence,
        save_all_chunks=False,
        response_language=response_language,
        max_questions=max_questions,
        question_offset=question_offset
    )
    summary = build_ai_score_summary(results)

    conn = get_connection()
    cur = conn.cursor()
    saved_run = save_domain_assessment_run(cur, domain_id, evidence_id, summary)
    conn.commit()
    cur.close()
    conn.close()

    return {
        "message": "Domain evidence analyzed successfully",
        "run_id": saved_run["id"],
        "model_name": AI_MODEL_NAME,
        "base_model_name": AI_BASE_MODEL_NAME,
        "domain": domain,
        "evidence_id": evidence_id,
        "analysis_limit": analysis_limit,
        "summary": summary,
        "questions": results,
        "created_at": str(saved_run["created_at"])
    }


@app.post("/api/ndmo/analyze-all-evidence")
def analyze_all_ndmo_evidence(
    evidence_id: int = Form(...),
    response_language: str = Form("English"),
    max_questions: int = Form(MAX_AI_QUESTIONS),
    question_offset: int = Form(0)
):
    create_ai_assessment_tables()

    evidence = get_evidence_by_id(evidence_id)
    if evidence is None:
        raise HTTPException(status_code=404, detail="Evidence not found")

    questions = get_all_ndmo_questions()
    if not questions:
        raise HTTPException(status_code=404, detail="No NDMO questions found")

    results, analysis_limit = analyze_questions_against_evidence(
        questions,
        evidence,
        save_all_chunks=False,
        response_language=response_language,
        max_questions=max_questions,
        question_offset=question_offset
    )

    domain_summaries = []
    conn = get_connection()
    cur = conn.cursor()

    domain_ids = sorted({
        result["domain_id"]
        for result in results
        if result.get("domain_id") is not None
    })

    for domain_id in domain_ids:
        domain = get_domain_by_id(domain_id)
        domain_results = [
            result for result in results
            if result.get("domain_id") == domain_id
        ]
        summary = build_ai_score_summary(domain_results)
        saved_run = save_domain_assessment_run(cur, domain_id, evidence_id, summary)

        domain_summaries.append({
            "run_id": saved_run["id"],
            "domain": domain,
            "summary": summary,
            "questions": domain_results,
            "created_at": str(saved_run["created_at"])
        })

    overall_summary = build_ai_score_summary(results)
    conn.commit()
    cur.close()
    conn.close()

    return {
        "message": "All NDMO evidence analyzed successfully",
        "model_name": AI_MODEL_NAME,
        "base_model_name": AI_BASE_MODEL_NAME,
        "evidence_id": evidence_id,
        "analysis_limit": analysis_limit,
        "overall_ndi_score": overall_summary["ndi_score"],
        "overall_summary": overall_summary,
        "domains": domain_summaries
    }


@app.post("/api/ndmo/ai-feedback")
def save_ndmo_ai_feedback(
    question_id: int = Form(...),
    evidence_id: int = Form(...),
    evidence_text: str = Form(...),
    corrected_answer: str = Form(...),
    predicted_answer: str = Form(""),
    evidence_location: str = Form(""),
    notes: str = Form("")
):
    create_ai_assessment_tables()

    normalized_answer = normalize_answer_label(corrected_answer)
    if normalized_answer not in ["yes", "partial", "no"]:
        raise HTTPException(
            status_code=400,
            detail="corrected_answer must be yes, partial, or no"
        )

    question = get_question_by_id(question_id)
    if question is None:
        raise HTTPException(status_code=404, detail="Question not found")

    evidence = get_evidence_by_id(evidence_id)
    if evidence is None:
        raise HTTPException(status_code=404, detail="Evidence not found")

    cleaned_evidence_text = normalize_training_text(evidence_text)
    cleaned_evidence_location = normalize_training_text(evidence_location)
    cleaned_predicted_answer = clean_text(predicted_answer).lower()
    cleaned_notes = normalize_training_text(notes)

    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT id, evidence_text, evidence_location, created_at
        FROM ndmo_ai_training_examples
        WHERE question_id = %s
          AND evidence_id = %s
          AND corrected_answer = %s;
    """, (
        question_id,
        evidence_id,
        normalized_answer
    ))

    existing_examples = cur.fetchall()
    for existing_example in existing_examples:
        if (
            normalize_training_text(existing_example["evidence_text"]) == cleaned_evidence_text
            and normalize_training_text(existing_example["evidence_location"]) == cleaned_evidence_location
        ):
            cur.close()
            conn.close()
            return {
                "message": "Training example already exists",
                "training_example_id": existing_example["id"],
                "duplicate": True,
                "model_name": AI_MODEL_NAME,
                "base_model_name": AI_BASE_MODEL_NAME,
                "question_id": question_id,
                "evidence_id": evidence_id,
                "corrected_answer": normalized_answer,
                "created_at": str(existing_example["created_at"])
            }

    cur.execute("""
        INSERT INTO ndmo_ai_training_examples
        (
            question_id,
            evidence_id,
            evidence_text,
            evidence_location,
            predicted_answer,
            corrected_answer,
            notes
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s)
        RETURNING id, created_at;
    """, (
        question_id,
        evidence_id,
        cleaned_evidence_text,
        cleaned_evidence_location,
        cleaned_predicted_answer,
        normalized_answer,
        cleaned_notes
    ))

    saved_example = cur.fetchone()
    conn.commit()
    cur.close()
    conn.close()

    return {
        "message": "AI feedback saved as a training example",
        "training_example_id": saved_example["id"],
        "duplicate": False,
        "model_name": AI_MODEL_NAME,
        "base_model_name": AI_BASE_MODEL_NAME,
        "question_id": question_id,
        "evidence_id": evidence_id,
        "corrected_answer": normalized_answer,
        "created_at": str(saved_example["created_at"])
    }


@app.get("/api/ndmo/training-dataset")
def get_ndmo_training_dataset():
    create_ai_assessment_tables()

    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            t.id,
            t.question_id,
            q.question_text_ar,
            q.question_text_en,
            t.evidence_id,
            t.evidence_text,
            t.evidence_location,
            t.predicted_answer,
            t.corrected_answer,
            t.notes,
            t.created_at
        FROM ndmo_ai_training_examples t
        JOIN ndmo_questions q ON q.id = t.question_id
        ORDER BY t.id;
    """)

    examples = cur.fetchall()
    cur.close()
    conn.close()

    return {
        "model_name": AI_MODEL_NAME,
        "base_model_name": AI_BASE_MODEL_NAME,
        "total_examples": len(examples),
        "examples": examples
    }


@app.get("/api/ndmo/model-info")
def get_ndmo_model_info():
    if AI_PROVIDER == "demo":
        current_stage = "Read-only synthetic demo with precomputed Qwen assessment results"
        next_step = "Run the full local profile to enable live Qwen and BGE-M3 inference."
    elif AI_PROVIDER == "ollama":
        current_stage = "Qwen via Ollama with multilingual semantic retrieval"
        next_step = "Collect reviewed examples for an NDMO-specific LoRA fine-tune."
    else:
        current_stage = "Rule-based analyzer with training data collection"
        next_step = "Collect corrected examples, then fine-tune the base model."

    return {
        "provider": AI_PROVIDER,
        "model_name": AI_MODEL_NAME,
        "base_model_name": AI_BASE_MODEL_NAME,
        "embedding_provider": EMBEDDING_PROVIDER,
        "embedding_model_name": EMBEDDING_MODEL_NAME if SEMANTIC_RETRIEVAL_ENABLED else None,
        "retrieval_mode": EVIDENCE_RETRIEVAL_MODE,
        "current_stage": current_stage,
        "labels": ["yes", "partial", "no"],
        "next_step": next_step
    }


@app.get("/api/ndmo/assessment-summary")
def get_ndmo_assessment_summary():
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        WITH latest_ai AS (
            SELECT DISTINCT ON (question_id)
                question_id,
                evidence_id,
                CASE LOWER(ai_answer)
                    WHEN 'yes' THEN 100
                    WHEN 'partial' THEN 50
                    ELSE 0
                END AS score
            FROM ndmo_ai_assessment_results
            ORDER BY question_id, created_at DESC, id DESC
        )
        SELECT
            d.id AS domain_id,
            d.name_ar AS domain_name_ar,
            d.name_en AS domain_name_en,
            COUNT(q.id) AS total_questions,
            COUNT(a.score) AS answered_questions,
            COALESCE(ROUND(AVG(COALESCE(a.score, 0)), 2), 0) AS domain_score,
            COUNT(DISTINCT a.evidence_id) AS uploaded_evidence
        FROM ndmo_domains d
        LEFT JOIN ndmo_questions q ON q.domain_id = d.id
        LEFT JOIN latest_ai a ON a.question_id = q.id
        GROUP BY d.id, d.name_ar, d.name_en, d.order_number
        HAVING COUNT(q.id) > 0
        ORDER BY d.order_number;
    """)

    domains_summary = cur.fetchall()

    cur.execute("""
        WITH latest_ai AS (
            SELECT DISTINCT ON (question_id)
                question_id,
                evidence_id,
                CASE LOWER(ai_answer)
                    WHEN 'yes' THEN 100
                    WHEN 'partial' THEN 50
                    ELSE 0
                END AS score
            FROM ndmo_ai_assessment_results
            ORDER BY question_id, created_at DESC, id DESC
        )
        SELECT
            COUNT(q.id) AS total_questions,
            COUNT(a.score) AS answered_questions,
            COALESCE(ROUND(AVG(COALESCE(a.score, 0)), 2), 0) AS overall_score,
            COUNT(DISTINCT a.evidence_id) AS uploaded_evidence
        FROM ndmo_questions q
        LEFT JOIN latest_ai a ON a.question_id = q.id;
    """)

    overall_summary = cur.fetchone()

    cur.close()
    conn.close()

    return {
        "overall_summary": overall_summary,
        "domains_summary": domains_summary
    }


@app.get("/api/ndmo/assessment-details/{domain_id}")
def get_ndmo_assessment_details(domain_id: int):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        WITH latest_ai AS (
            SELECT DISTINCT ON (question_id)
                id, question_id, evidence_id, ai_answer, confidence, reason,
                evidence_text, evidence_location, created_at
            FROM ndmo_ai_assessment_results
            ORDER BY question_id, created_at DESC, id DESC
        ), latest_manual AS (
            SELECT DISTINCT ON (question_id)
                id, question_id, answer_value, score, notes, created_at
            FROM ndmo_answers
            ORDER BY question_id, created_at DESC, id DESC
        )
        SELECT
            q.id, q.question_code, q.question_text_ar, q.question_text_en,
            q.evidence_required, q.max_score, q.domain_id,
            d.name_ar AS domain_name_ar, d.name_en AS domain_name_en,
            m.answer_value AS manual_answer, m.score AS manual_score,
            m.notes AS manual_notes, m.created_at AS manual_created_at,
            a.id AS result_id, a.ai_answer, a.confidence, a.reason,
            a.evidence_text, a.evidence_location, a.created_at AS analyzed_at,
            a.evidence_id, ev.file_name AS evidence_file,
            ev.evidence_type, ev.status AS evidence_status,
            COALESCE(m.answer_value, a.ai_answer) AS effective_answer,
            COALESCE(m.score,
                CASE LOWER(a.ai_answer) WHEN 'yes' THEN 100 WHEN 'partial' THEN 50 WHEN 'no' THEN 0 END
            ) AS effective_score,
            CASE WHEN m.id IS NOT NULL THEN 'manual' WHEN a.id IS NOT NULL THEN 'evidence_analysis' ELSE NULL END AS decision_source
        FROM ndmo_questions q
        JOIN ndmo_domains d ON d.id = q.domain_id
        LEFT JOIN latest_manual m ON m.question_id = q.id
        LEFT JOIN latest_ai a ON a.question_id = q.id
        LEFT JOIN ndmo_evidence ev ON ev.id = a.evidence_id
        WHERE q.domain_id = %s
        ORDER BY q.id
    """, (domain_id,))
    questions = cur.fetchall()
    cur.close()
    conn.close()
    if not questions:
        raise HTTPException(status_code=404, detail="Domain not found or has no requirements")
    return {"domain_id": domain_id, "questions": questions}
